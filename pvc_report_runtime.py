from __future__ import annotations

import base64
import html
import io
import json
import math
import re
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any, Callable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, Spacer, Table, TableStyle
from reportlab.platypus.tableofcontents import TableOfContents

import app as core
import institutional_export as institutional
import pdf_progress_runtime
import period_policy_runtime
import period_unified_runtime
import report_quality
import report_structure
import requirements_store
from db import connection, rows_to_dicts, utcnow
from workflow_rules import prerequisite_state


PASS_GRADE = 7.0
FORMULA_TOLERANCE = 0.05
WRITTEN_WEIGHT = 0.70
DEFENSE_WEIGHT = 0.30
SOURCE_LABEL = "Base institucional de Requisitos y Base de resultados PVC procesadas por Informtit."

PRACTICAL_CRITERIA = (
    ("Diseño", "diseno", 2.5),
    ("Construcción", "construccion", 2.5),
    ("Funcionamiento", "funcionamiento", 2.5),
    ("Aplicación", "aplicacion", 2.5),
)
DEFENSE_CRITERIA = (
    ("Sustento del marco teórico", "sustentomarcoteorico", 2.0),
    ("Sustento de la propuesta", "sustentopropuesta", 2.0),
    ("Utilización de recursos", "utilizacionrecursos", 2.0),
    ("Solvencia en preguntas", "solvenciapreguntas", 4.0),
)

_BASE_BUILD_PDF: Callable[[int], Path] | None = None
_BASE_PROJECT_SUMMARY: Callable[[int], dict[str, Any]] | None = None
_INSTALLED = False


def _fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.casefold())


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).strip().split())


def _identification(value: Any) -> str:
    text = _text(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return re.sub(r"\D", "", text)


def _number(value: Any) -> float | None:
    text = _text(value)
    if not text or text.casefold() in {"null", "none", "nan", "-", "—", "n/a"}:
        return None
    try:
        if "," in text and "." not in text:
            text = text.replace(",", ".")
        return float(text)
    except (TypeError, ValueError):
        return None


def _is_pvc(report_id: int) -> bool:
    period_policy_runtime.ensure_schema()
    with connection() as conn:
        row = conn.execute(
            "SELECT report_type, period FROM reports WHERE id=?",
            (int(report_id),),
        ).fetchone()
    if not row:
        return False
    kind = str(row["report_type"] or period_policy_runtime.classify_period(row["period"]))
    return kind == "pvc"


def ensure_schema() -> None:
    period_policy_runtime.ensure_schema()
    requirements_store.ensure_requirements_schema()
    with connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS pvc_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER NOT NULL,
                filename TEXT DEFAULT '',
                source_period TEXT DEFAULT '',
                total_rows INTEGER DEFAULT 0,
                matched_rows INTEGER DEFAULT 0,
                unmatched_rows INTEGER DEFAULT 0,
                formula_warnings INTEGER DEFAULT 0,
                imported_at TEXT NOT NULL,
                FOREIGN KEY(report_id) REFERENCES reports(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_pvc_imports_report
                ON pvc_imports(report_id, imported_at);

            CREATE TABLE IF NOT EXISTS pvc_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_id INTEGER NOT NULL,
                source_row INTEGER DEFAULT 0,
                identification TEXT DEFAULT '',
                source_name TEXT DEFAULT '',
                official_name TEXT DEFAULT '',
                career_code TEXT DEFAULT '',
                career_name TEXT DEFAULT '',
                campus TEXT DEFAULT '',
                source_period TEXT DEFAULT '',
                work_type TEXT DEFAULT '',
                act_number TEXT DEFAULT '',
                act_date TEXT DEFAULT '',
                tutor_name TEXT DEFAULT '',
                reader_name TEXT DEFAULT '',
                vocal_1 TEXT DEFAULT '',
                vocal_2 TEXT DEFAULT '',
                vocal_3 TEXT DEFAULT '',
                tutor_grade REAL,
                reader_grade REAL,
                written_source REAL,
                written_calculated REAL,
                defense_source REAL,
                final_source REAL,
                final_calculated REAL,
                accumulated_grade REAL,
                practical_json TEXT DEFAULT '[]',
                defense_json TEXT DEFAULT '[]',
                requirements_complete INTEGER DEFAULT 0,
                missing_requirements_json TEXT DEFAULT '[]',
                match_status TEXT DEFAULT 'UNMATCHED',
                formula_status TEXT DEFAULT 'INCOMPLETE',
                final_status TEXT DEFAULT 'NO EVALUADO',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(report_id) REFERENCES reports(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_pvc_records_report
                ON pvc_records(report_id);
            CREATE INDEX IF NOT EXISTS idx_pvc_records_identification
                ON pvc_records(report_id, identification);
            """
        )


def _decode_data_url(data_url: str) -> bytes:
    if "," not in str(data_url or ""):
        raise ValueError("El archivo PVC no fue enviado correctamente.")
    header, encoded = str(data_url).split(",", 1)
    if ";base64" not in header:
        raise ValueError("La Base PVC debe enviarse como archivo codificado.")
    try:
        return base64.b64decode(encoded)
    except Exception as exc:
        raise ValueError("No se pudo decodificar la Base PVC.") from exc


def _headers(values: tuple[Any, ...]) -> dict[str, list[int]]:
    result: dict[str, list[int]] = defaultdict(list)
    for index, value in enumerate(values):
        key = _fold(value)
        if key:
            result[key].append(index)
    return dict(result)


def _cell(row: tuple[Any, ...], indexes: dict[str, list[int]], key: str, occurrence: int = 0) -> Any:
    positions = indexes.get(_fold(key), [])
    if occurrence >= len(positions):
        return None
    position = positions[occurrence]
    return row[position] if position < len(row) else None


def _criterion_rows(
    row: tuple[Any, ...],
    indexes: dict[str, list[int]],
    definitions: tuple[tuple[str, str, float], ...],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for label, key, maximum in definitions:
        result.append(
            {
                "criterion": label,
                "max_score": maximum,
                "vocal_1": _number(_cell(row, indexes, f"{key}vocal1")),
                "vocal_2": _number(_cell(row, indexes, f"{key}vocal2")),
                "vocal_3": _number(_cell(row, indexes, f"{key}vocal3")),
            }
        )
    return result


def _formula_state(
    written_source: float | None,
    written_calculated: float | None,
    final_source: float | None,
    final_calculated: float | None,
) -> str:
    warnings = []
    if written_source is not None and written_calculated is not None:
        warnings.append(abs(written_source - written_calculated) > FORMULA_TOLERANCE)
    if final_source is not None and final_calculated is not None:
        warnings.append(abs(final_source - final_calculated) > FORMULA_TOLERANCE)
    if any(warnings):
        return "WARNING"
    if written_calculated is None or final_source is None or final_calculated is None:
        return "INCOMPLETE"
    return "OK"


def _final_status(
    final_source: float | None,
    written: float | None,
    defense: float | None,
) -> str:
    if final_source is None:
        return "NO EVALUADO"
    if final_source == 0 and written is None and defense is None:
        return "NO EVALUADO"
    return "APROBADO" if final_source >= PASS_GRADE else "REPROBADO"


def parse_pvc_workbook(data: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("No se pudo abrir la Base PVC. Utilice un archivo .xlsx válido.") from exc
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ValueError("La Base PVC está vacía.") from exc
    indexes = _headers(header)

    required = (
        "nombre_estudiante",
        "identificacion_estudiante",
        "evaluacionTutor",
        "evaluacionLector",
        "nota_defensa_oral",
        "notaTrabajoTitulacion",
    )
    missing = [key for key in required if _fold(key) not in indexes]
    if missing:
        raise ValueError("La Base PVC no contiene las columnas requeridas: " + ", ".join(missing))

    result: list[dict[str, Any]] = []
    for source_row, row in enumerate(rows, start=2):
        identification = _identification(_cell(row, indexes, "identificacion_estudiante"))
        source_name = _text(_cell(row, indexes, "nombre_estudiante"))
        if not identification and not source_name:
            continue

        tutor = _number(_cell(row, indexes, "evaluacionTutor"))
        reader = _number(_cell(row, indexes, "evaluacionLector"))
        written_calc = round((tutor + reader) / 2, 2) if tutor is not None and reader is not None else None
        written_source = _number(_cell(row, indexes, "promedio_trabajo_escrito", 0))
        # El archivo suministrado contiene el encabezado promedio_trabajo_escrito
        # dos veces. El primer valor es la fuente principal; el segundo solo sirve
        # como respaldo cuando el primero está vacío.
        if written_source is None:
            written_source = _number(_cell(row, indexes, "promedio_trabajo_escrito", 1))
        defense_source = _number(_cell(row, indexes, "nota_defensa_oral"))
        final_source = _number(_cell(row, indexes, "notaTrabajoTitulacion"))
        effective_written = written_calc if written_calc is not None else written_source
        final_calc = (
            round((effective_written * WRITTEN_WEIGHT) + (defense_source * DEFENSE_WEIGHT), 2)
            if effective_written is not None and defense_source is not None
            else None
        )
        result.append(
            {
                "source_row": source_row,
                "identification": identification,
                "source_name": source_name,
                "source_period": _text(_cell(row, indexes, "periodo_academico")),
                "work_type": _text(_cell(row, indexes, "trabajoTitulacion")),
                "act_number": _text(_cell(row, indexes, "numeroActaGrado")),
                "act_date": _text(_cell(row, indexes, "fechaActaGrado")),
                "tutor_name": _text(_cell(row, indexes, "nombre_tutor")),
                "reader_name": _text(_cell(row, indexes, "nombre_lector")),
                "vocal_1": _text(_cell(row, indexes, "nombre_vocal1")),
                "vocal_2": _text(_cell(row, indexes, "nombre_vocal2")),
                "vocal_3": _text(_cell(row, indexes, "nombre_vocal3")),
                "tutor_grade": tutor,
                "reader_grade": reader,
                "written_source": written_source,
                "written_calculated": written_calc,
                "defense_source": defense_source,
                "final_source": final_source,
                "final_calculated": final_calc,
                "accumulated_grade": _number(_cell(row, indexes, "notaPromedioAcumulado")),
                "practical": _criterion_rows(row, indexes, PRACTICAL_CRITERIA),
                "defense": _criterion_rows(row, indexes, DEFENSE_CRITERIA),
                "formula_status": _formula_state(
                    written_source, written_calc, final_source, final_calc
                ),
                "final_status": _final_status(final_source, effective_written, defense_source),
            }
        )
    if not result:
        raise ValueError("No se detectaron estudiantes en la Base PVC.")

    seen_identifications: dict[str, int] = {}
    duplicate_rows: list[tuple[str, int, int]] = []
    for item in result:
        identification = str(item.get("identification") or "").strip()
        if not identification:
            continue
        previous = seen_identifications.get(identification)
        if previous is not None:
            duplicate_rows.append(
                (identification, previous, int(item.get("source_row") or 0))
            )
        else:
            seen_identifications[identification] = int(item.get("source_row") or 0)
    if duplicate_rows:
        detail = "; ".join(
            f"{identification} (filas {first} y {second})"
            for identification, first, second in duplicate_rows[:8]
        )
        if len(duplicate_rows) > 8:
            detail += f"; y {len(duplicate_rows) - 8} duplicado(s) adicional(es)"
        raise ValueError(
            "La Base PVC contiene cédulas duplicadas. Corrija el archivo antes de importarlo: "
            + detail
        )
    return result


def _requirements_by_id(report_id: int) -> dict[str, list[dict[str, Any]]]:
    requirements_store.ensure_requirements_schema()
    with connection() as conn:
        rows = rows_to_dicts(
            conn.execute(
                "SELECT * FROM requirements_students WHERE report_id=? ORDER BY id",
                (report_id,),
            ).fetchall()
        )
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = _identification(row.get("identification"))
        if key:
            groups[key].append(row)
    return dict(groups)


def import_pvc_results(report_id: int, data_url: str, filename: str) -> dict[str, Any]:
    ensure_schema()
    if not _is_pvc(report_id):
        raise ValueError("La Base PVC solo puede cargarse en un informe identificado como PVC.")
    if not str(filename or "").lower().endswith(".xlsx"):
        raise ValueError("La Base de resultados PVC debe cargarse en formato .xlsx.")

    records = parse_pvc_workbook(_decode_data_url(data_url))
    requirements = _requirements_by_id(report_id)
    now = utcnow()
    matched = 0
    unmatched = 0
    formula_warnings = 0
    source_periods: Counter[str] = Counter()

    with connection() as conn:
        conn.execute("DELETE FROM pvc_records WHERE report_id=?", (report_id,))
        for record in records:
            source_periods[record["source_period"] or "Sin período"] += 1
            candidates = requirements.get(record["identification"], [])
            if len(candidates) == 1:
                official = candidates[0]
                req_state = prerequisite_state(official)
                match_status = "MATCHED"
                matched += 1
                official_name = _text(official.get("full_name"))
                career_code = _text(official.get("career_code"))
                career_name = _text(official.get("career_name"))
                campus = _text(official.get("campus"))
                requirements_complete = int(bool(req_state["complete"]))
                missing_requirements = list(req_state["missing"])
            elif len(candidates) > 1:
                match_status = "AMBIGUOUS_REQUIREMENTS"
                unmatched += 1
                official_name = career_code = career_name = campus = ""
                requirements_complete = 0
                missing_requirements = []
            else:
                match_status = "OUT_OF_REQUIREMENTS"
                unmatched += 1
                official_name = career_code = career_name = campus = ""
                requirements_complete = 0
                missing_requirements = []

            formula_warnings += int(record["formula_status"] == "WARNING")
            conn.execute(
                """
                INSERT INTO pvc_records
                (report_id, source_row, identification, source_name, official_name,
                 career_code, career_name, campus, source_period, work_type,
                 act_number, act_date, tutor_name, reader_name, vocal_1, vocal_2,
                 vocal_3, tutor_grade, reader_grade, written_source,
                 written_calculated, defense_source, final_source, final_calculated,
                 accumulated_grade, practical_json, defense_json,
                 requirements_complete, missing_requirements_json, match_status,
                 formula_status, final_status, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    report_id,
                    record["source_row"],
                    record["identification"],
                    record["source_name"],
                    official_name,
                    career_code,
                    career_name,
                    campus,
                    record["source_period"],
                    record["work_type"],
                    record["act_number"],
                    record["act_date"],
                    record["tutor_name"],
                    record["reader_name"],
                    record["vocal_1"],
                    record["vocal_2"],
                    record["vocal_3"],
                    record["tutor_grade"],
                    record["reader_grade"],
                    record["written_source"],
                    record["written_calculated"],
                    record["defense_source"],
                    record["final_source"],
                    record["final_calculated"],
                    record["accumulated_grade"],
                    json.dumps(record["practical"], ensure_ascii=False),
                    json.dumps(record["defense"], ensure_ascii=False),
                    requirements_complete,
                    json.dumps(missing_requirements, ensure_ascii=False),
                    match_status,
                    record["formula_status"],
                    record["final_status"],
                    now,
                    now,
                ),
            )

        dominant_period = source_periods.most_common(1)[0][0] if source_periods else ""
        conn.execute(
            """
            INSERT INTO pvc_imports
            (report_id, filename, source_period, total_rows, matched_rows,
             unmatched_rows, formula_warnings, imported_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report_id,
                _text(filename),
                dominant_period if dominant_period != "Sin período" else "",
                len(records),
                matched,
                unmatched,
                formula_warnings,
                now,
            ),
        )

    return {
        "ok": True,
        "report_id": report_id,
        "filename": _text(filename),
        "records": len(records),
        "matched": matched,
        "unmatched": unmatched,
        "formula_warnings": formula_warnings,
        "source_periods": dict(source_periods),
    }


def _effective_written(row: dict[str, Any]) -> float | None:
    return _number(row.get("written_calculated")) if _number(row.get("written_calculated")) is not None else _number(row.get("written_source"))


def _effective_final(row: dict[str, Any]) -> float | None:
    return _number(row.get("final_source")) if _number(row.get("final_source")) is not None else _number(row.get("final_calculated"))


def _safe_mean(values: list[float | None]) -> float | None:
    valid = [float(value) for value in values if value is not None and math.isfinite(float(value))]
    return round(mean(valid), 2) if valid else None


def _pct(part: int | float, total: int | float) -> float:
    return round((float(part) * 100.0 / float(total)), 2) if total else 0.0


def _record_dict(row: dict[str, Any]) -> dict[str, Any]:
    item = dict(row)
    try:
        item["practical"] = json.loads(item.get("practical_json") or "[]")
    except (TypeError, json.JSONDecodeError):
        item["practical"] = []
    try:
        item["defense"] = json.loads(item.get("defense_json") or "[]")
    except (TypeError, json.JSONDecodeError):
        item["defense"] = []
    try:
        item["missing_requirements"] = json.loads(item.get("missing_requirements_json") or "[]")
    except (TypeError, json.JSONDecodeError):
        item["missing_requirements"] = []
    item["written"] = _effective_written(item)
    item["final_grade"] = _effective_final(item)
    item["display_name"] = item.get("official_name") or item.get("source_name") or "Sin nombre"
    return item


def get_pvc_summary(report_id: int, include_records: bool = True) -> dict[str, Any]:
    ensure_schema()
    if not _is_pvc(report_id):
        raise ValueError("El informe solicitado no es PVC.")

    roster = requirements_store.get_report_roster(report_id)
    requirements = list(roster.get("students") or [])
    with connection() as conn:
        raw_records = rows_to_dicts(
            conn.execute(
                "SELECT * FROM pvc_records WHERE report_id=? ORDER BY career_name, official_name, source_name, id",
                (report_id,),
            ).fetchall()
        )
        latest_import = conn.execute(
            "SELECT * FROM pvc_imports WHERE report_id=? ORDER BY id DESC LIMIT 1",
            (report_id,),
        ).fetchone()
        report = conn.execute("SELECT * FROM reports WHERE id=?", (report_id,)).fetchone()

    records = [_record_dict(row) for row in raw_records]
    eligible = [row for row in requirements if prerequisite_state(row)["complete"]]
    evaluated = [row for row in records if row["final_status"] in {"APROBADO", "REPROBADO"}]
    approved = [row for row in evaluated if row["final_status"] == "APROBADO"]
    failed = [row for row in evaluated if row["final_status"] == "REPROBADO"]
    not_evaluated = [row for row in records if row["final_status"] == "NO EVALUADO"]
    matched = [row for row in records if row["match_status"] == "MATCHED"]
    unmatched = [row for row in records if row["match_status"] != "MATCHED"]
    formula_warnings = [row for row in records if row["formula_status"] == "WARNING"]

    requirements_by_career: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in requirements:
        requirements_by_career[_text(row.get("career_name")) or "Sin carrera"].append(row)
    records_by_career: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in matched:
        records_by_career[_text(row.get("career_name")) or "Sin carrera"].append(row)

    career_names = sorted(set(requirements_by_career) | set(records_by_career))
    by_career = []
    for career in career_names:
        req_rows = requirements_by_career.get(career, [])
        pvc_rows = records_by_career.get(career, [])
        eligible_rows = [row for row in req_rows if prerequisite_state(row)["complete"]]
        career_eval = [row for row in pvc_rows if row["final_status"] in {"APROBADO", "REPROBADO"}]
        career_approved = [row for row in career_eval if row["final_status"] == "APROBADO"]
        by_career.append(
            {
                "career": career,
                "requirements": len(req_rows),
                "eligible": len(eligible_rows),
                "pvc": len(pvc_rows),
                "evaluated": len(career_eval),
                "approved": len(career_approved),
                "failed": len([row for row in career_eval if row["final_status"] == "REPROBADO"]),
                "not_evaluated": len([row for row in pvc_rows if row["final_status"] == "NO EVALUADO"]),
                "approval_pct": _pct(len(career_approved), len(career_eval)),
                "written_average": _safe_mean([row["written"] for row in career_eval]),
                "defense_average": _safe_mean([_number(row.get("defense_source")) for row in career_eval]),
                "final_average": _safe_mean([row["final_grade"] for row in career_eval]),
            }
        )

    requirement_failures: Counter[str] = Counter()
    for row in requirements:
        state = prerequisite_state(row)
        for label in state["missing"]:
            requirement_failures[str(label)] += 1

    campus_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in matched:
        campus_groups[_text(row.get("campus")) or "Sin sede"].append(row)
    by_campus = []
    for campus, campus_rows in sorted(campus_groups.items()):
        campus_eval = [row for row in campus_rows if row["final_status"] in {"APROBADO", "REPROBADO"}]
        campus_approved = [row for row in campus_eval if row["final_status"] == "APROBADO"]
        by_campus.append(
            {
                "campus": campus,
                "records": len(campus_rows),
                "evaluated": len(campus_eval),
                "approval_pct": _pct(len(campus_approved), len(campus_eval)),
                "final_average": _safe_mean([row["final_grade"] for row in campus_eval]),
            }
        )

    source_periods = Counter(_text(row.get("source_period")) or "Sin período" for row in records)
    work_types = Counter(_text(row.get("work_type")) or "Sin denominación" for row in records)

    summary = {
        "requirements_total": len(requirements),
        "eligible": len(eligible),
        "pvc_total": len(records),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "evaluated": len(evaluated),
        "approved": len(approved),
        "failed": len(failed),
        "not_evaluated": len(not_evaluated),
        "approval_pct": _pct(len(approved), len(evaluated)),
        "written_average": _safe_mean([row["written"] for row in evaluated]),
        "defense_average": _safe_mean([_number(row.get("defense_source")) for row in evaluated]),
        "final_average": _safe_mean([row["final_grade"] for row in evaluated]),
        "formula_warnings": len(formula_warnings),
    }
    return {
        "ok": True,
        "report": dict(report) if report else {},
        "summary": summary,
        "by_career": by_career,
        "requirements_failures": [
            {"label": label, "count": count}
            for label, count in requirement_failures.most_common()
        ],
        "by_campus": by_campus,
        "source_periods": dict(source_periods),
        "work_types": dict(work_types),
        "latest_import": dict(latest_import) if latest_import else None,
        "unmatched_records": unmatched[:100],
        "formula_warning_records": formula_warnings[:100],
        "records": records if include_records else [],
    }


def _criterion_summary(records: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    definitions = PRACTICAL_CRITERIA if key == "practical" else DEFENSE_CRITERIA
    buckets: dict[str, list[float]] = defaultdict(list)
    for record in records:
        for row in record.get(key, []):
            maximum = _number(row.get("max_score"))
            if not maximum:
                continue
            values = [_number(row.get(vocal)) for vocal in ("vocal_1", "vocal_2", "vocal_3")]
            valid = [value for value in values if value is not None]
            if valid:
                normalized = mean(valid) / maximum * 10.0
                buckets[_text(row.get("criterion"))].append(normalized)
    result = []
    for label, _, _ in definitions:
        values = buckets.get(label, [])
        if values:
            result.append({"criterion": label, "score": round(mean(values), 2)})
    return result


def _vocal_summary(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[str, list[float]] = defaultdict(list)
    for record in records:
        for vocal in ("vocal_1", "vocal_2", "vocal_3"):
            practical_values = [
                _number(row.get(vocal)) for row in record.get("practical", [])
            ]
            defense_values = [
                _number(row.get(vocal)) for row in record.get("defense", [])
            ]
            if practical_values and defense_values and all(value is not None for value in practical_values + defense_values):
                practical_total = sum(float(value) for value in practical_values if value is not None)
                defense_total = sum(float(value) for value in defense_values if value is not None)
                buckets[vocal].append((practical_total + defense_total) / 2.0)
    labels = {"vocal_1": "Vocal 1", "vocal_2": "Vocal 2", "vocal_3": "Vocal 3"}
    return [
        {"vocal": labels[key], "average": round(mean(values), 2)}
        for key, values in buckets.items()
        if values
    ]


def _historical_pvc(current_report_id: int) -> list[dict[str, Any]]:
    ensure_schema()
    with connection() as conn:
        reports = rows_to_dicts(
            conn.execute(
                """
                SELECT id, period FROM reports
                WHERE report_type='pvc'
                ORDER BY id
                """
            ).fetchall()
        )
    result = []
    for report in reports:
        report_id = int(report["id"])
        try:
            data = get_pvc_summary(report_id, include_records=False)
        except Exception:
            continue
        if data["summary"]["evaluated"] <= 0:
            continue
        result.append(
            {
                "report_id": report_id,
                "period": report.get("period") or str(report_id),
                "approval_pct": data["summary"]["approval_pct"],
                "final_average": data["summary"]["final_average"],
                "current": report_id == current_report_id,
            }
        )
    return result


def pvc_audit(report_id: int) -> dict[str, Any]:
    data = get_pvc_summary(report_id, include_records=False)
    summary = data["summary"]
    report_period = _text(data["report"].get("period"))
    source_periods = [key for key in data["source_periods"] if key and key != "Sin período"]
    controls: list[dict[str, Any]] = []

    def add(name: str, status: str, detail: str, blocking: bool = False) -> None:
        controls.append(
            {"name": name, "status": status, "detail": detail, "blocking": blocking}
        )

    add(
        "Población maestra de Requisitos",
        "ok" if summary["requirements_total"] else "error",
        f"{summary['requirements_total']} estudiantes registrados en Requisitos."
        if summary["requirements_total"]
        else "No existe población de Requisitos para el PVC.",
        blocking=not bool(summary["requirements_total"]),
    )
    add(
        "Base de resultados PVC",
        "ok" if summary["pvc_total"] else "error",
        f"{summary['pvc_total']} registros de resultados PVC cargados."
        if summary["pvc_total"]
        else "Cargue la Base de resultados PVC (.xlsx) antes de generar el informe.",
        blocking=not bool(summary["pvc_total"]),
    )
    add(
        "Conciliación por cédula",
        "ok" if summary["unmatched"] == 0 else "warning",
        f"{summary['matched']} coincidencias exactas y {summary['unmatched']} registros sin identidad oficial confirmada.",
    )
    add(
        "Validación de fórmula 70/30",
        "ok" if summary["formula_warnings"] == 0 else "warning",
        f"{summary['formula_warnings']} registros presentan diferencia superior a {FORMULA_TOLERANCE:.2f} entre la nota fuente y el cálculo 70 % trabajo escrito + 30 % defensa.",
    )
    if source_periods:
        normalized_report = period_policy_runtime.canonical_period_id(report_period)
        normalized_sources = {
            period_policy_runtime.canonical_period_id(value) for value in source_periods
        }
        mismatch = bool(normalized_report and any(value and value != normalized_report for value in normalized_sources))
        add(
            "Período de la Base PVC",
            "warning" if mismatch else "ok",
            "Período(s) detectado(s) en la base: " + ", ".join(source_periods)
            + (". Revise la diferencia con el período configurado en el informe." if mismatch else "."),
        )
    blocking = [item for item in controls if item["status"] == "error" and item["blocking"]]
    warning_count = sum(item["status"] == "warning" for item in controls)
    can_generate = not blocking
    return {
        "ok": True,
        "state": "APTO PARA EMITIR" if can_generate and warning_count == 0 else "BORRADOR",
        "mode": "pvc",
        "final_ready": bool(can_generate and warning_count == 0),
        "can_generate_pdf": can_generate,
        "controls": controls,
        "blocking_errors": blocking,
        "reconciliation_label": "Conciliación PVC",
        "reconciliation": {
            "imported": summary["pvc_total"],
            "included": summary["matched"],
            "excluded": summary["unmatched"],
            "reasons": {
                "Sin coincidencia en Requisitos": summary["unmatched"],
                "Advertencias de fórmula": summary["formula_warnings"],
            },
        },
    }


# ---------------------------------------------------------------------------
# PDF helpers: la regla contexto -> elemento -> análisis se centraliza aquí.


def _styles() -> Any:
    styles = report_quality._pdf_styles()
    if "PvcCentered" not in styles:
        styles.add(
            ParagraphStyle(
                "PvcCentered",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=9,
                leading=12,
                alignment=TA_CENTER,
                spaceAfter=5,
            )
        )
    if "PvcSmall" not in styles:
        styles.add(
            ParagraphStyle(
                "PvcSmall",
                parent=styles["BodyText"],
                fontName="Helvetica",
                fontSize=7.2,
                leading=9,
                alignment=TA_JUSTIFY,
                spaceAfter=3,
            )
        )
    return styles


def _body(story: list[Any], styles: Any, text: str) -> None:
    if str(text or "").strip():
        story.append(
            Paragraph(
                html.escape(" ".join(str(text).split())),
                styles["BodyJustified"],
            )
        )


def _bullet(story: list[Any], styles: Any, text: str) -> None:
    story.append(
        Paragraph("• " + html.escape(" ".join(str(text).split())), styles["BulletIndented"])
    )


def _heading(story: list[Any], context: report_quality.ExportContext, styles: Any, level: int, title: str) -> None:
    if level == 1:
        if context.major_started:
            story.append(PageBreak())
        context.major_started = True
    story.append(
        Paragraph(
            html.escape(context.heading(level, title)),
            styles[f"Heading{level}"],
        )
    )


def _table(
    headers: list[str],
    rows: list[list[Any]],
    widths: list[float],
    font_size: float = 7.0,
) -> Table:
    prepared = []
    for row in [headers] + rows:
        prepared.append(
            [
                Paragraph(
                    html.escape("—" if value is None or value == "" else str(value)),
                    ParagraphStyle(
                        "PvcTableCell",
                        fontName="Helvetica-Bold" if row is headers else "Helvetica",
                        fontSize=font_size,
                        leading=font_size + 1.5,
                    ),
                )
                for value in row
            ]
        )
    table = Table(prepared, repeatRows=1, colWidths=widths)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#244a73")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _add_table_block(
    story: list[Any],
    context: report_quality.ExportContext,
    styles: Any,
    registry: dict[str, list[str]],
    title: str,
    context_text: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[float],
    analysis_text: str,
    note: str = SOURCE_LABEL,
    font_size: float = 7.0,
) -> None:
    _body(story, styles, context_text)
    caption = context.table_caption(title)
    registry["tables"].append(caption)
    story.append(Paragraph(html.escape(caption), styles["FigureCaption"]))
    story.append(_table(headers, rows, widths, font_size=font_size))
    story.append(Spacer(1, 0.08 * cm))
    story.append(Paragraph(html.escape("Nota. " + note), styles["FigureCaption"]))
    _body(story, styles, analysis_text)
    story.append(Spacer(1, 0.16 * cm))


def _add_figure_block(
    story: list[Any],
    context: report_quality.ExportContext,
    styles: Any,
    registry: dict[str, list[str]],
    title: str,
    context_text: str,
    image_path: Path,
    analysis_text: str,
    note: str = SOURCE_LABEL,
    width_cm: float = 16.2,
    height_cm: float = 9.3,
) -> None:
    _body(story, styles, context_text)
    caption = context.figure_caption(title)
    registry["figures"].append(caption)
    story.append(Paragraph(html.escape(caption), styles["FigureCaption"]))
    story.append(institutional.fit_image(image_path, width_cm * cm, height_cm * cm))
    story.append(Spacer(1, 0.06 * cm))
    story.append(Paragraph(html.escape("Nota. " + note), styles["FigureCaption"]))
    _body(story, styles, analysis_text)
    story.append(Spacer(1, 0.18 * cm))


def _chart_path(temp_paths: list[Path]) -> Path:
    handle = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    handle.close()
    path = Path(handle.name)
    temp_paths.append(path)
    return path


def _save(fig: Any, path: Path) -> None:
    fig.tight_layout()
    fig.savefig(path, dpi=155, bbox_inches="tight")
    plt.close(fig)


def _bar_chart(labels: list[str], values: list[float], title: str, ylabel: str, path: Path, horizontal: bool = False) -> None:
    fig, ax = plt.subplots(figsize=(9, 5.2))
    if horizontal:
        ax.barh(labels, values)
        ax.set_xlabel(ylabel)
        ax.invert_yaxis()
        for index, value in enumerate(values):
            ax.text(value, index, f" {value:.2f}" if isinstance(value, float) else f" {value}", va="center", fontsize=8)
    else:
        ax.bar(labels, values)
        ax.set_ylabel(ylabel)
        ax.tick_params(axis="x", rotation=30)
        for index, value in enumerate(values):
            ax.text(index, value, f"{value:.2f}" if isinstance(value, float) else str(value), ha="center", va="bottom", fontsize=8)
    ax.set_title(title)
    ax.grid(axis="y" if not horizontal else "x", alpha=0.18)
    _save(fig, path)


def _grouped_chart(labels: list[str], first: list[float], second: list[float], first_label: str, second_label: str, title: str, path: Path) -> None:
    fig, ax = plt.subplots(figsize=(10, 5.4))
    positions = list(range(len(labels)))
    width = 0.38
    ax.bar([x - width / 2 for x in positions], first, width, label=first_label)
    ax.bar([x + width / 2 for x in positions], second, width, label=second_label)
    ax.set_xticks(positions)
    ax.set_xticklabels(labels, rotation=30, ha="right")
    ax.set_ylim(bottom=0)
    ax.set_title(title)
    ax.legend()
    ax.grid(axis="y", alpha=0.18)
    _save(fig, path)


def _funnel_chart(summary: dict[str, Any], path: Path) -> None:
    labels = ["Requisitos", "Habilitados", "Base PVC", "Evaluados", "Aprobados"]
    values = [
        summary["requirements_total"],
        summary["eligible"],
        summary["pvc_total"],
        summary["evaluated"],
        summary["approved"],
    ]
    fig, ax = plt.subplots(figsize=(9, 5))
    positions = list(range(len(labels)))
    ax.barh(positions, values)
    ax.set_yticks(positions)
    ax.set_yticklabels(labels)
    ax.invert_yaxis()
    ax.set_xlabel("Número de estudiantes")
    ax.set_title("Flujo cuantitativo del proceso PVC")
    for index, value in enumerate(values):
        ax.text(value, index, f" {value}", va="center", fontsize=9)
    ax.grid(axis="x", alpha=0.18)
    _save(fig, path)


def _stacked_requirements(by_career: list[dict[str, Any]], path: Path) -> None:
    labels = [row["career"] for row in by_career]
    eligible = [row["eligible"] for row in by_career]
    pending = [max(0, row["requirements"] - row["eligible"]) for row in by_career]
    fig, ax = plt.subplots(figsize=(10, max(5, len(labels) * 0.42)))
    ax.barh(labels, eligible, label="Habilitados")
    ax.barh(labels, pending, left=eligible, label="No habilitados / pendientes")
    ax.invert_yaxis()
    ax.set_xlabel("Estudiantes")
    ax.set_title("Habilitación de requisitos por carrera")
    ax.legend()
    ax.grid(axis="x", alpha=0.18)
    _save(fig, path)


def _pareto(failures: list[dict[str, Any]], path: Path) -> None:
    labels = [row["label"] for row in failures]
    values = [int(row["count"]) for row in failures]
    total = sum(values) or 1
    cumulative = []
    running = 0
    for value in values:
        running += value
        cumulative.append(running * 100 / total)
    fig, ax = plt.subplots(figsize=(10, 5.2))
    ax.bar(labels, values)
    ax.set_ylabel("Casos")
    ax.tick_params(axis="x", rotation=30)
    ax2 = ax.twinx()
    ax2.plot(labels, cumulative, marker="o")
    ax2.set_ylabel("Porcentaje acumulado")
    ax2.set_ylim(0, 105)
    ax.set_title("Pareto de requisitos pendientes")
    _save(fig, path)


def _status_chart(summary: dict[str, Any], path: Path) -> None:
    labels = ["Aprobados", "Reprobados", "No evaluados"]
    values = [summary["approved"], summary["failed"], summary["not_evaluated"]]
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.pie(values, labels=labels, autopct=lambda pct: f"{pct:.1f}%" if pct > 0 else "")
    ax.set_title("Estado final de los registros PVC")
    _save(fig, path)


def _histogram(records: list[dict[str, Any]], path: Path) -> None:
    grades = [row["final_grade"] for row in records if row["final_grade"] is not None and row["final_status"] in {"APROBADO", "REPROBADO"}]
    fig, ax = plt.subplots(figsize=(8.8, 5))
    ax.hist(grades, bins=[0, 7, 8, 9, 9.5, 10.01], edgecolor="black")
    ax.set_xticks([0, 7, 8, 9, 9.5, 10])
    ax.set_xlabel("Calificación final")
    ax.set_ylabel("Estudiantes")
    ax.set_title("Distribución de calificaciones finales PVC")
    ax.grid(axis="y", alpha=0.18)
    _save(fig, path)


def _scatter_tutor_reader(records: list[dict[str, Any]], path: Path) -> None:
    pairs = [
        (_number(row.get("tutor_grade")), _number(row.get("reader_grade")))
        for row in records
    ]
    pairs = [(x, y) for x, y in pairs if x is not None and y is not None]
    fig, ax = plt.subplots(figsize=(7, 6))
    if pairs:
        ax.scatter([x for x, _ in pairs], [y for _, y in pairs], alpha=0.7)
    ax.plot([0, 10], [0, 10], linestyle="--", linewidth=1)
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 10)
    ax.set_xlabel("Calificación del tutor")
    ax.set_ylabel("Calificación del lector")
    ax.set_title("Concordancia descriptiva entre tutor y lector")
    ax.grid(alpha=0.18)
    _save(fig, path)


def _heatmap(by_career: list[dict[str, Any]], path: Path) -> None:
    labels = [row["career"] for row in by_career if row["pvc"]]
    rows = [row for row in by_career if row["pvc"]]
    metrics = ["written_average", "defense_average", "final_average", "approval_pct"]
    titles = ["Trabajo escrito", "Defensa", "Nota final", "Aprobación %"]
    matrix = []
    for row in rows:
        matrix.append([
            row.get("written_average") or 0,
            row.get("defense_average") or 0,
            row.get("final_average") or 0,
            (row.get("approval_pct") or 0) / 10.0,
        ])
    fig, ax = plt.subplots(figsize=(9.5, max(4.5, len(labels) * 0.42)))
    image = ax.imshow(matrix, aspect="auto")
    ax.set_yticks(range(len(labels)))
    ax.set_yticklabels(labels)
    ax.set_xticks(range(len(titles)))
    ax.set_xticklabels(titles, rotation=20, ha="right")
    for i, row in enumerate(matrix):
        for j, value in enumerate(row):
            display = by_career[[r["career"] for r in by_career].index(labels[i])].get(metrics[j])
            ax.text(j, i, "—" if display is None else f"{display:.2f}", ha="center", va="center", fontsize=7)
    ax.set_title("Mapa de calor de desempeño por carrera")
    fig.colorbar(image, ax=ax, fraction=0.025, pad=0.02)
    _save(fig, path)


def _flow_chart(path: Path) -> None:
    steps = [
        "Requisitos",
        "Inducción PVC",
        "Metodología y entregas",
        "Tutoría",
        "Artículo final",
        "Antiplagio y rúbrica",
        "Defensa",
        "Resultado y cierre",
    ]
    fig, ax = plt.subplots(figsize=(11, 6.2))
    ax.axis("off")
    coordinates = [
        (0.08, 0.78), (0.32, 0.78), (0.56, 0.78), (0.80, 0.78),
        (0.80, 0.28), (0.56, 0.28), (0.32, 0.28), (0.08, 0.28),
    ]
    for index, (label, (x, y)) in enumerate(zip(steps, coordinates)):
        ax.text(x, y, label, ha="center", va="center", fontsize=9,
                bbox={"boxstyle": "round,pad=0.5", "fc": "white", "ec": "black"})
        if index < len(coordinates) - 1:
            nx, ny = coordinates[index + 1]
            ax.annotate("", xy=(nx, ny), xytext=(x, y),
                        arrowprops={"arrowstyle": "->", "lw": 1.2})
    ax.set_title("Flujo institucional del proceso PVC – Artículo Científico")
    _save(fig, path)


def _traceability_chart(path: Path) -> None:
    fig, ax = plt.subplots(figsize=(10, 5.8))
    ax.axis("off")
    boxes = [
        (0.12, 0.65, "REQUISITOS\nIdentidad · carrera · sede\nrequisitos habilitantes"),
        (0.50, 0.65, "CONCILIACIÓN\nCédula como llave principal\nvalidación de identidad"),
        (0.88, 0.65, "BASE PVC\nTutor · lector · tribunal\nrúbrica · defensa · nota"),
        (0.50, 0.22, "INFORME PVC\nIndicadores · tablas · gráficos\nconclusiones · mejora"),
    ]
    for x, y, text in boxes:
        ax.text(x, y, text, ha="center", va="center", fontsize=9,
                bbox={"boxstyle": "round,pad=0.6", "fc": "white", "ec": "black"})
    ax.annotate("", xy=(0.40, 0.65), xytext=(0.22, 0.65), arrowprops={"arrowstyle": "->"})
    ax.annotate("", xy=(0.78, 0.65), xytext=(0.60, 0.65), arrowprops={"arrowstyle": "->"})
    ax.annotate("", xy=(0.50, 0.31), xytext=(0.50, 0.55), arrowprops={"arrowstyle": "->"})
    ax.annotate("", xy=(0.56, 0.31), xytext=(0.82, 0.58), arrowprops={"arrowstyle": "->"})
    ax.set_title("Trazabilidad de fuentes del informe PVC")
    _save(fig, path)


def _ishikawa(data: dict[str, Any], path: Path) -> None:
    summary = data["summary"]
    failures = data["requirements_failures"][:3]
    evaluated = [row for row in data["records"] if row["final_status"] in {"APROBADO", "REPROBADO"}]
    low_written = sum(row["written"] is not None and row["written"] < PASS_GRADE for row in evaluated)
    low_defense = sum(_number(row.get("defense_source")) is not None and float(row["defense_source"]) < PASS_GRADE for row in evaluated)
    branches = [
        ("Requisitos", [f"{row['label']}: {row['count']}" for row in failures] or ["Sin concentración registrada"]),
        ("Trabajo escrito", [f"Calificación < 7: {low_written}"]),
        ("Defensa", [f"Calificación < 7: {low_defense}"]),
        ("Datos", [f"Sin conciliar: {summary['unmatched']}", f"Fórmula por revisar: {summary['formula_warnings']}"]),
        ("Seguimiento", [f"No evaluados: {summary['not_evaluated']}"]),
        ("Resultado", [f"Reprobados evaluados: {summary['failed']}"]),
    ]
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.axis("off")
    ax.plot([0.12, 0.84], [0.5, 0.5], linewidth=2)
    ax.annotate("", xy=(0.92, 0.5), xytext=(0.84, 0.5), arrowprops={"arrowstyle": "->", "lw": 2})
    ax.text(0.93, 0.5, "No culminación o\ndesempeño insuficiente", va="center", fontsize=9)
    positions = [(0.25, 0.78), (0.48, 0.78), (0.70, 0.78), (0.25, 0.20), (0.48, 0.20), (0.70, 0.20)]
    for (title, items), (x, y) in zip(branches, positions):
        target_y = 0.5
        ax.plot([x, x + 0.10], [y, target_y], linewidth=1.2)
        ax.text(x - 0.02, y, title + "\n" + "\n".join(items), ha="center", va="center", fontsize=8,
                bbox={"boxstyle": "round,pad=0.35", "fc": "white", "ec": "grey"})
    ax.set_title("Ishikawa de factores asociados y aspectos por verificar")
    _save(fig, path)


def _priority_matrix(by_career: list[dict[str, Any]], path: Path) -> None:
    rows = [row for row in by_career if row["requirements"] and row["evaluated"]]
    fig, ax = plt.subplots(figsize=(8.5, 6))
    for row in rows:
        pending = 100.0 - _pct(row["eligible"], row["requirements"])
        non_approval = 100.0 - row["approval_pct"]
        ax.scatter([pending], [non_approval])
        ax.annotate(row["career"], (pending, non_approval), xytext=(4, 4), textcoords="offset points", fontsize=7)
    ax.set_xlabel("No habilitación en Requisitos (%)")
    ax.set_ylabel("No aprobación entre evaluados (%)")
    ax.set_title("Matriz descriptiva de priorización para acciones de mejora")
    ax.grid(alpha=0.2)
    _save(fig, path)


def _history_chart(history: list[dict[str, Any]], path: Path) -> None:
    labels = [row["period"] for row in history]
    values = [row["approval_pct"] for row in history]
    fig, ax = plt.subplots(figsize=(9.5, 5))
    ax.plot(labels, values, marker="o")
    ax.set_ylim(0, 105)
    ax.set_ylabel("Aprobación (%)")
    ax.tick_params(axis="x", rotation=25)
    ax.set_title("Evolución histórica de la aprobación PVC")
    ax.grid(alpha=0.2)
    _save(fig, path)


def _cover(report: dict[str, Any], styles: Any) -> list[Any]:
    title = ParagraphStyle(
        "PvcCover",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
    )
    story: list[Any] = [
        Spacer(1, 3.4 * cm),
        Paragraph("Informe Final Del Proceso De Titulación", title),
        Spacer(1, 0.25 * cm),
        Paragraph("PVC – Modalidad Artículo Científico", title),
        Paragraph(html.escape(str(report.get("period") or "")), title),
        Spacer(1, 6.3 * cm),
    ]
    data = [[
        institutional.signature_items(
            report, "ELABORADO POR:", institutional.SIG_PREPARED,
            str(report.get("prepared_by") or ""), str(report.get("prepared_role") or ""), styles
        ),
        institutional.signature_items(
            report, "REVISADO POR:", institutional.SIG_REVIEWED,
            str(report.get("reviewed_by") or ""), str(report.get("reviewed_role") or ""), styles
        ),
        institutional.signature_items(
            report, "APROBADO POR:", institutional.SIG_APPROVED,
            str(report.get("approved_by") or ""), str(report.get("approved_role") or ""), styles
        ),
    ]]
    table = Table(data, colWidths=[5.55 * cm] * 3)
    table.setStyle(
        TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 5),
        ])
    )
    story += [table, PageBreak()]
    return story


def _fmt(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "—"
    return f"{number:.2f}".replace(".", ",")


def _best_worst(by_career: list[dict[str, Any]], key: str) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    rows = [row for row in by_career if row.get(key) is not None and row.get("evaluated", 0) > 0]
    if not rows:
        return None, None
    return max(rows, key=lambda row: float(row[key])), min(rows, key=lambda row: float(row[key]))


def _pvc_count(report_id: int | None) -> int:
    if not report_id:
        return 0
    ensure_schema()
    with connection() as conn:
        return int(conn.execute(
            "SELECT COUNT(*) FROM pvc_records WHERE report_id=?",
            (int(report_id),),
        ).fetchone()[0])


def build_pvc_pdf(report_id: int) -> Path:
    if not _is_pvc(report_id):
        if _BASE_BUILD_PDF is None:
            raise RuntimeError("Generador PDF base no disponible.")
        return Path(_BASE_BUILD_PDF(report_id))

    audit = pdf_progress_runtime.consume_preflight(report_id, "pvc") or pvc_audit(report_id)
    if not audit["can_generate_pdf"]:
        detail = "; ".join(item["detail"] for item in audit["blocking_errors"])
        raise ValueError("No se puede generar el PDF PVC: " + detail)

    pdf_progress_runtime._set_progress(8, "Preparando informe PVC", "Consolidando Requisitos y Base de resultados PVC.")
    data = get_pvc_summary(report_id, include_records=True)
    summary = data["summary"]
    report = report_quality._report_data(report_id)
    report["report_type"] = "pvc"
    report["modality"] = "presencial"
    styles = _styles()
    context = report_quality.ExportContext.create()
    registry: dict[str, list[str]] = {"tables": [], "figures": []}
    body_story: list[Any] = []
    temp_paths: list[Path] = []

    # 1. Resumen ejecutivo
    _heading(body_story, context, styles, 1, "Resumen ejecutivo")
    _body(
        body_story, styles,
        "El presente resumen sintetiza la trayectoria de la cohorte PVC desde la población registrada en Requisitos hasta los resultados de la modalidad de Artículo Científico. Los indicadores se calculan únicamente con los registros disponibles y distinguen a los estudiantes no evaluados de los reprobados."
    )
    _add_table_block(
        body_story, context, styles, registry,
        "Indicadores ejecutivos del proceso PVC",
        "Para establecer una visión general del período, se presentan los principales indicadores de población, habilitación, evaluación y desempeño final. La tasa de aprobación utiliza como denominador exclusivamente a los estudiantes con evaluación final válida.",
        ["Indicador", "Resultado"],
        [
            ["Estudiantes en Requisitos", summary["requirements_total"]],
            ["Habilitados por requisitos", summary["eligible"]],
            ["Registros en Base PVC", summary["pvc_total"]],
            ["Coincidencias exactas por cédula", summary["matched"]],
            ["Evaluados", summary["evaluated"]],
            ["Aprobados", summary["approved"]],
            ["Reprobados", summary["failed"]],
            ["No evaluados", summary["not_evaluated"]],
            ["Aprobación entre evaluados", f"{summary['approval_pct']:.2f} %"],
            ["Promedio trabajo escrito", _fmt(summary["written_average"])],
            ["Promedio defensa", _fmt(summary["defense_average"])],
            ["Promedio final", _fmt(summary["final_average"])],
        ],
        [10.8 * cm, 5.5 * cm],
        f"El período registra {summary['requirements_total']} estudiantes en la población maestra y {summary['pvc_total']} registros en la Base PVC. De {summary['evaluated']} estudiantes evaluados, {summary['approved']} aprobaron, equivalente al {summary['approval_pct']:.2f} %. Los {summary['not_evaluated']} casos sin evaluación se mantienen separados del cálculo de aprobación.",
    )
    funnel = _chart_path(temp_paths)
    _funnel_chart(summary, funnel)
    _add_figure_block(
        body_story, context, styles, registry,
        "Embudo de eficiencia terminal del proceso PVC",
        "El embudo permite observar el número de estudiantes disponible en cada etapa cuantificable: población de Requisitos, habilitación, presencia en la Base PVC, evaluación y aprobación. La diferencia entre etapas representa casos que deben interpretarse según el estado administrativo o académico disponible y no equivale automáticamente a reprobación.",
        funnel,
        f"La mayor diferencia entre etapas debe revisarse comparando Requisitos y Base PVC. Actualmente existen {summary['eligible']} estudiantes habilitados y {summary['pvc_total']} registros PVC; la conciliación identificó {summary['unmatched']} registros que no pudieron vincularse de forma única con la población maestra.",
    )

    # 2. Aspectos generales
    pdf_progress_runtime._set_progress(16, "Redactando marco institucional PVC", "Incorporando objetivos, alcance y lineamientos del Artículo Científico.")
    _heading(body_story, context, styles, 1, "Aspectos generales")
    _heading(body_story, context, styles, 2, "Introducción")
    _body(
        body_story, styles,
        "El Informe Final del Proceso de Titulación PVC – Modalidad Artículo Científico consolida la información académica y administrativa del período, desde la verificación de requisitos habilitantes hasta la evaluación del artículo, la defensa oral y el cierre del proceso. El documento integra datos de Requisitos y de la Base de resultados PVC, manteniendo trazabilidad por cédula y preservando la fuente oficial de identidad, carrera y sede."
    )
    _heading(body_story, context, styles, 2, "Objetivo general")
    _body(
        body_story, styles,
        "Evaluar la implementación del proceso de titulación PVC en la modalidad de Artículo Científico mediante el análisis de sus fases, población, resultados y desempeño por carreras, con el fin de aportar evidencia para la calidad y la mejora continua."
    )
    _heading(body_story, context, styles, 2, "Objetivos específicos")
    for item in (
        "Describir las fases metodológicas y procedimentales que conforman el proceso PVC.",
        "Registrar y analizar los indicadores de habilitación, evaluación, aprobación y calidad de datos.",
        "Analizar el desempeño del trabajo escrito, la defensa oral y los criterios registrados en la rúbrica.",
        "Comparar los resultados entre carreras y, cuando exista información histórica suficiente, entre cohortes PVC.",
        "Formular conclusiones y acciones de mejora sustentadas en los hallazgos del período.",
    ):
        _bullet(body_story, styles, item)
    _heading(body_story, context, styles, 2, "Alcance")
    _body(
        body_story, styles,
        "El alcance comprende exclusivamente el proceso PVC de Artículo Científico del período configurado. Incluye la población registrada en Requisitos, los registros de la Base PVC, el acompañamiento académico, las evaluaciones disponibles, la defensa y el resultado final. Los procesos de Examen Complexivo y otras rutas de titulación no forman parte de este informe."
    )

    _heading(body_story, context, styles, 1, "Marco normativo e institucional")
    _body(
        body_story, styles,
        "El informe se desarrolla en el marco de la normativa nacional de educación superior y de los instrumentos institucionales aplicables a titulación, calidad, investigación y trazabilidad documental. La aplicación concreta de cada disposición debe corresponder a la versión normativa vigente registrada por la institución para el período."
    )
    _add_table_block(
        body_story, context, styles, registry,
        "Referentes normativos e institucionales del PVC",
        "A continuación se identifican los referentes normativos que enmarcan el proceso. La tabla resume su función dentro del informe y evita reproducir extensamente disposiciones legales que deben consultarse en sus fuentes oficiales.",
        ["Referente", "Aplicación en el informe PVC"],
        [
            ["Constitución de la República del Ecuador", "Derecho, calidad y finalidad de la educación superior."],
            ["Ley Orgánica de Educación Superior", "Egreso, titulación y garantía de procesos académicos."],
            ["Reglamento de Régimen Académico", "Organización de las opciones y procesos de titulación."],
            ["Reglamento / lineamientos institucionales de UTET", "Procedimiento, evaluación, evidencias y cierre del PVC."],
            ["Modelo de aseguramiento de la calidad", "Trazabilidad, indicadores, análisis y mejora continua."],
        ],
        [6.2 * cm, 10.1 * cm],
        "Los referentes se articulan alrededor de tres necesidades del informe: demostrar el cumplimiento del proceso, mantener evidencia verificable y utilizar los resultados para la mejora continua. Informtit no sustituye la validación jurídica de la versión normativa vigente.",
    )

    # 3. Metodología
    _heading(body_story, context, styles, 1, "Metodología del proceso PVC")
    _body(
        body_story, styles,
        "La metodología del PVC se estructura como una secuencia de acompañamiento, construcción del artículo, validación y evaluación. Los documentos institucionales de referencia describen clases de metodología, entregas secuenciales, tutoría con docente investigador, revisión antiplagio, rúbrica institucional, tribunal de tres miembros, defensa oral y, cuando corresponda, mecanismos de recuperación."
    )
    flow = _chart_path(temp_paths)
    _flow_chart(flow)
    _add_figure_block(
        body_story, context, styles, registry,
        "Flujo institucional del proceso PVC",
        "El siguiente flujo representa el orden lógico de las etapas que estructuran el PVC. Su propósito es mostrar cómo la habilitación administrativa antecede al trabajo académico y cómo las evidencias del artículo y la defensa convergen en el resultado final.",
        flow,
        "El flujo evidencia que la nota final es el resultado de un proceso previo de habilitación, acompañamiento y evaluación. Por ello, un estudiante sin registro final no debe clasificarse automáticamente como reprobado: su situación debe mantenerse como no evaluada o pendiente según la evidencia disponible.",
        note="Elaboración de Informtit con base en la estructura institucional de los informes PVC suministrados.",
        height_cm=10.0,
    )
    trace = _chart_path(temp_paths)
    _traceability_chart(trace)
    _add_figure_block(
        body_story, context, styles, registry,
        "Mapa de trazabilidad de las fuentes de información",
        "La elaboración automática del informe requiere separar la fuente maestra de identidad de la fuente de resultados. Requisitos conserva la información oficial del estudiante, mientras que la Base PVC aporta tutor, lector, tribunal, rúbricas y calificaciones.",
        trace,
        f"La conciliación por cédula logró {summary['matched']} coincidencias exactas de {summary['pvc_total']} registros PVC. Los {summary['unmatched']} casos restantes se mantienen fuera de los indicadores por carrera hasta confirmar su identidad, evitando atribuir resultados a una carrera incorrecta.",
        note=SOURCE_LABEL,
        height_cm=9.2,
    )
    _add_table_block(
        body_story, context, styles, registry,
        "Ponderación y componentes de evaluación del PVC",
        "La Base PVC permite verificar el promedio del trabajo escrito a partir de las evaluaciones de tutor y lector. La nota final se valida mediante la ponderación institucional utilizada en los informes PVC de referencia: 70 % para el trabajo escrito y 30 % para la defensa oral.",
        ["Componente", "Cálculo / fuente", "Ponderación final"],
        [
            ["Trabajo escrito", "Promedio de tutor y lector", "70 %"],
            ["Defensa oral", "Nota de defensa registrada en la Base PVC", "30 %"],
            ["Nota final PVC", "(Trabajo escrito × 0,70) + (Defensa × 0,30)", "100 %"],
        ],
        [4.2 * cm, 8.3 * cm, 3.8 * cm],
        f"Informtit recalcula la fórmula sin reemplazar silenciosamente el dato fuente. En este período se detectaron {summary['formula_warnings']} registros con diferencias superiores a {FORMULA_TOLERANCE:.2f}; estos casos se presentan como advertencias de calidad de datos.",
    )

    # 4. Población y calidad
    pdf_progress_runtime._set_progress(30, "Analizando población y requisitos", "Calculando habilitación, conciliación y causas de pendientes.")
    _heading(body_story, context, styles, 1, "Población y calidad de datos")
    career_rows = [
        [
            row["career"], row["requirements"], row["eligible"], row["pvc"],
            row["evaluated"], row["approved"], row["failed"], row["not_evaluated"],
        ]
        for row in data["by_career"]
    ]
    _add_table_block(
        body_story, context, styles, registry,
        "Población y avance del proceso por carrera",
        "Para interpretar correctamente la eficiencia terminal, la población se presenta por etapas. Requisitos corresponde a la base maestra; habilitados son quienes cumplen los requisitos definidos; Base PVC refleja los resultados importados y Evaluados corresponde a registros con una nota final interpretable.",
        ["Carrera", "Req.", "Hab.", "PVC", "Eval.", "Aprob.", "Reprob.", "No eval."],
        career_rows,
        [5.1 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.4 * cm, 1.6 * cm],
        f"La cohorte registra {summary['requirements_total']} estudiantes en Requisitos, de los cuales {summary['eligible']} se encuentran habilitados. La Base PVC contiene {summary['pvc_total']} registros y {summary['evaluated']} evaluaciones válidas. Esta separación permite localizar diferencias entre habilitación, participación y evaluación.",
        font_size=6.2,
    )
    if data["by_career"]:
        stacked = _chart_path(temp_paths)
        _stacked_requirements(data["by_career"], stacked)
        _add_figure_block(
            body_story, context, styles, registry,
            "Habilitación de requisitos por carrera",
            "El gráfico compara, para cada carrera, la cantidad de estudiantes habilitados frente a quienes mantienen requisitos pendientes o información insuficiente. Esta lectura se realiza antes de analizar las calificaciones PVC.",
            stacked,
            f"A nivel institucional, {summary['eligible']} de {summary['requirements_total']} estudiantes cumplen los requisitos habilitantes. Las diferencias entre carreras deben priorizarse según el número absoluto de casos y no únicamente por porcentaje.",
            height_cm=max(8.0, min(15.0, len(data["by_career"]) * 0.55)),
        )
    if data["requirements_failures"]:
        pareto = _chart_path(temp_paths)
        _pareto(data["requirements_failures"][:10], pareto)
        top_failure = data["requirements_failures"][0]
        _add_figure_block(
            body_story, context, styles, registry,
            "Pareto de requisitos pendientes",
            "El Pareto ordena los requisitos faltantes de mayor a menor frecuencia y añade el porcentaje acumulado. Su finalidad es identificar dónde se concentra la mayor cantidad de pendientes para orientar acciones de seguimiento.",
            pareto,
            f"El requisito con mayor frecuencia de casos pendientes es {top_failure['label']} con {top_failure['count']} registros. Este resultado representa una concentración administrativa observable y no implica por sí mismo una causa de reprobación académica.",
        )

    quality_rows = [
        ["Registros Base PVC", summary["pvc_total"]],
        ["Coincidencias exactas por cédula", summary["matched"]],
        ["Sin coincidencia única", summary["unmatched"]],
        ["Advertencias de fórmula", summary["formula_warnings"]],
        ["Períodos detectados en la fuente", "; ".join(data["source_periods"].keys()) or "—"],
        ["Denominación(es) del trabajo", "; ".join(data["work_types"].keys()) or "—"],
    ]
    _add_table_block(
        body_story, context, styles, registry,
        "Controles de calidad y conciliación de la Base PVC",
        "Antes de utilizar las calificaciones para análisis institucional, se verifican identidad, período, denominación del trabajo y consistencia de las fórmulas. Los casos sin coincidencia se conservan como evidencia, pero no se incorporan a comparaciones por carrera.",
        ["Control", "Resultado"],
        quality_rows,
        [8.5 * cm, 7.8 * cm],
        f"La calidad de datos presenta {summary['unmatched']} casos sin coincidencia oficial y {summary['formula_warnings']} advertencias de fórmula. Estos hallazgos deben resolverse en la fuente o mediante conciliación documentada antes del cierre definitivo del período.",
    )

    # 5. Resultados
    pdf_progress_runtime._set_progress(46, "Generando resultados y gráficos PVC", "Construyendo análisis por carrera, componentes y evaluadores.")
    _heading(body_story, context, styles, 1, "Resultados del proceso PVC")
    active_careers = [row for row in data["by_career"] if row["pvc"] > 0]
    if active_careers:
        pop = _chart_path(temp_paths)
        _bar_chart(
            [row["career"] for row in active_careers],
            [float(row["pvc"]) for row in active_careers],
            "Registros PVC por carrera", "Estudiantes", pop, horizontal=True
        )
        max_pop = max(active_careers, key=lambda row: row["pvc"])
        _add_figure_block(
            body_story, context, styles, registry,
            "Distribución de registros PVC por carrera",
            "La cantidad de registros por carrera contextualiza los porcentajes y promedios posteriores. Una tasa calculada sobre una población pequeña no debe interpretarse con el mismo peso descriptivo que una carrera con mayor número de estudiantes.",
            pop,
            f"La carrera con mayor número de registros PVC conciliados es {max_pop['career']} con {max_pop['pvc']} estudiantes. Las comparaciones posteriores deben considerar esta diferencia de tamaño de cohorte.",
            height_cm=max(7.5, min(14.0, len(active_careers) * 0.5)),
        )

        participation = _chart_path(temp_paths)
        _grouped_chart(
            [row["career"] for row in active_careers],
            [float(row["eligible"]) for row in active_careers],
            [float(row["pvc"]) for row in active_careers],
            "Habilitados", "Registros PVC",
            "Habilitación frente a participación efectiva", participation
        )
        _add_figure_block(
            body_story, context, styles, registry,
            "Habilitados frente a registros de resultados PVC",
            "La comparación entre habilitados y registros en la Base PVC permite detectar diferencias de cobertura entre la población que cumplió requisitos y la población con resultados cargados.",
            participation,
            f"En el total institucional se identifican {summary['eligible']} habilitados y {summary['pvc_total']} registros PVC. Una diferencia entre ambas cifras debe revisarse individualmente antes de atribuirla a abandono, ausencia o reprobación.",
            height_cm=9.5,
        )

    status = _chart_path(temp_paths)
    _status_chart(summary, status)
    _add_figure_block(
        body_story, context, styles, registry,
        "Estado final de los registros PVC",
        "El estado final diferencia aprobados, reprobados y no evaluados. Esta separación evita que la ausencia de una calificación sea interpretada como una nota de cero.",
        status,
        f"De {summary['pvc_total']} registros PVC, {summary['approved']} constan como aprobados, {summary['failed']} como reprobados y {summary['not_evaluated']} como no evaluados. La aprobación entre los {summary['evaluated']} evaluados es {summary['approval_pct']:.2f} %.",
        width_cm=12.5,
        height_cm=8.5,
    )

    if active_careers:
        approval = _chart_path(temp_paths)
        _bar_chart(
            [row["career"] for row in active_careers],
            [float(row["approval_pct"]) for row in active_careers],
            "Aprobación por carrera", "Aprobación (%)", approval, horizontal=True
        )
        best, worst = _best_worst(active_careers, "approval_pct")
        analysis = (
            f"Entre las carreras con estudiantes evaluados, {best['career']} presenta la mayor tasa descriptiva de aprobación ({best['approval_pct']:.2f} %) y {worst['career']} la menor ({worst['approval_pct']:.2f} %). "
            "La lectura debe considerar el número de evaluados y los casos no evaluados de cada carrera."
            if best and worst else
            "No existen suficientes evaluaciones por carrera para establecer una comparación descriptiva de aprobación."
        )
        _add_figure_block(
            body_story, context, styles, registry,
            "Porcentaje de aprobación por carrera",
            "El porcentaje de aprobación se calcula sobre estudiantes con evaluación final válida de cada carrera. Los no evaluados se reportan de forma separada y no forman parte del denominador.",
            approval,
            analysis,
            height_cm=max(8.0, min(15.0, len(active_careers) * 0.55)),
        )

        grade = _chart_path(temp_paths)
        grade_rows = [row for row in active_careers if row["final_average"] is not None]
        if grade_rows:
            _bar_chart(
                [row["career"] for row in grade_rows],
                [float(row["final_average"]) for row in grade_rows],
                "Promedio final PVC por carrera", "Calificación (0-10)", grade, horizontal=True
            )
            best_grade, worst_grade = _best_worst(grade_rows, "final_average")
            _add_figure_block(
                body_story, context, styles, registry,
                "Promedio final PVC por carrera",
                "El promedio final complementa la tasa de aprobación al mostrar el nivel de desempeño de quienes cuentan con una calificación válida. Se presenta en escala de 0 a 10.",
                grade,
                f"El mayor promedio final corresponde a {best_grade['career']} ({best_grade['final_average']:.2f}) y el menor a {worst_grade['career']} ({worst_grade['final_average']:.2f}). Esta comparación es descriptiva y debe interpretarse junto con el tamaño de la cohorte.",
                height_cm=max(8.0, min(15.0, len(grade_rows) * 0.55)),
            )

        written_defense = [row for row in active_careers if row["written_average"] is not None and row["defense_average"] is not None]
        if written_defense:
            wd = _chart_path(temp_paths)
            _grouped_chart(
                [row["career"] for row in written_defense],
                [float(row["written_average"]) for row in written_defense],
                [float(row["defense_average"]) for row in written_defense],
                "Trabajo escrito", "Defensa",
                "Trabajo escrito frente a defensa por carrera", wd
            )
            gaps = [
                (abs(float(row["written_average"]) - float(row["defense_average"])), row)
                for row in written_defense
            ]
            gap, gap_row = max(gaps, key=lambda item: item[0])
            _add_figure_block(
                body_story, context, styles, registry,
                "Comparación entre trabajo escrito y defensa por carrera",
                "La comparación de los dos componentes permite identificar si el desempeño relativo de una carrera es mayor en el documento escrito o en la presentación y argumentación oral.",
                wd,
                f"La mayor diferencia absoluta entre ambos componentes se observa en {gap_row['career']} con {gap:.2f} puntos. La brecha es un indicador descriptivo para orientar acompañamiento y no demuestra por sí sola una deficiencia metodológica u oral.",
                height_cm=9.5,
            )

    hist = _chart_path(temp_paths)
    _histogram(data["records"], hist)
    _add_figure_block(
        body_story, context, styles, registry,
        "Distribución de calificaciones finales",
        "La distribución permite observar la concentración de resultados dentro de intervalos de calificación y complementa los promedios institucionales.",
        hist,
        f"El promedio final de los registros evaluados es {_fmt(summary['final_average'])}. La distribución permite identificar si la cohorte se concentra cerca del umbral de aprobación o en rangos superiores, sin convertir los registros no evaluados en notas cero.",
    )

    scatter = _chart_path(temp_paths)
    _scatter_tutor_reader(data["records"], scatter)
    tutor_reader_pairs = [
        abs(float(row["tutor_grade"]) - float(row["reader_grade"]))
        for row in data["records"]
        if _number(row.get("tutor_grade")) is not None and _number(row.get("reader_grade")) is not None
    ]
    mean_gap = round(mean(tutor_reader_pairs), 2) if tutor_reader_pairs else None
    _add_figure_block(
        body_story, context, styles, registry,
        "Concordancia descriptiva entre calificaciones de tutor y lector",
        "Cada punto representa a un estudiante con calificación de tutor y lector. La diagonal indica igualdad entre ambas valoraciones y permite visualizar diferencias sin asumir que una de las dos evaluaciones sea incorrecta.",
        scatter,
        f"La diferencia absoluta promedio entre tutor y lector es {_fmt(mean_gap)} puntos." if mean_gap is not None else "No existen suficientes pares de calificaciones de tutor y lector para estimar una diferencia promedio.",
        width_cm=12.5,
        height_cm=9.5,
    )

    evaluated_records = [row for row in data["records"] if row["final_status"] in {"APROBADO", "REPROBADO"}]
    practical_summary = _criterion_summary(evaluated_records, "practical")
    defense_summary = _criterion_summary(evaluated_records, "defense")
    if practical_summary:
        practical_path = _chart_path(temp_paths)
        _bar_chart(
            [row["criterion"] for row in practical_summary],
            [float(row["score"]) for row in practical_summary],
            "Desempeño normalizado en evaluación práctica", "Índice (0-10)", practical_path, horizontal=True
        )
        weakest = min(practical_summary, key=lambda row: row["score"])
        strongest = max(practical_summary, key=lambda row: row["score"])
        _add_figure_block(
            body_story, context, styles, registry,
            "Desempeño por criterio de evaluación práctica",
            "Para hacer comparables los criterios de la rúbrica, los puntajes se normalizan a una escala de 0 a 10 a partir del máximo definido en la Base PVC. El gráfico consolida las valoraciones de los tres vocales.",
            practical_path,
            f"El criterio con mayor desempeño relativo es {strongest['criterion']} ({strongest['score']:.2f}) y el de menor desempeño es {weakest['criterion']} ({weakest['score']:.2f}). Este último puede considerarse un aspecto para seguimiento académico.",
        )
    if defense_summary:
        defense_path = _chart_path(temp_paths)
        _bar_chart(
            [row["criterion"] for row in defense_summary],
            [float(row["score"]) for row in defense_summary],
            "Desempeño normalizado en la defensa", "Índice (0-10)", defense_path, horizontal=True
        )
        weakest = min(defense_summary, key=lambda row: row["score"])
        strongest = max(defense_summary, key=lambda row: row["score"])
        _add_figure_block(
            body_story, context, styles, registry,
            "Desempeño por criterio de defensa",
            "Los criterios de defensa se normalizan a escala de 0 a 10 para comparar componentes con máximos distintos. Se consolidan las valoraciones disponibles de los tres vocales.",
            defense_path,
            f"El criterio con mayor desempeño relativo es {strongest['criterion']} ({strongest['score']:.2f}) y el menor es {weakest['criterion']} ({weakest['score']:.2f}). El resultado orienta el tipo de preparación oral que conviene reforzar.",
        )

    vocal_summary = _vocal_summary(evaluated_records)
    if len(vocal_summary) >= 2:
        vocal_path = _chart_path(temp_paths)
        _bar_chart(
            [row["vocal"] for row in vocal_summary],
            [float(row["average"]) for row in vocal_summary],
            "Promedio consolidado por posición de vocal", "Calificación (0-10)", vocal_path
        )
        spread = max(row["average"] for row in vocal_summary) - min(row["average"] for row in vocal_summary)
        _add_figure_block(
            body_story, context, styles, registry,
            "Consistencia descriptiva entre los tres vocales",
            "El gráfico compara el promedio consolidado de las calificaciones asignadas desde cada posición del tribunal. Su objetivo es observar diferencias sistemáticas de valoración, no evaluar individualmente a los docentes.",
            vocal_path,
            f"La diferencia entre el promedio más alto y el más bajo de las posiciones de vocal es {spread:.2f} puntos. Diferencias pequeñas sugieren comportamiento agregado similar; valores mayores justifican revisar la aplicación de la rúbrica.",
            width_cm=12.5,
            height_cm=8.2,
        )

    if len(active_careers) >= 2:
        heat = _chart_path(temp_paths)
        _heatmap(active_careers, heat)
        _add_figure_block(
            body_story, context, styles, registry,
            "Mapa de calor del desempeño por carrera y componente",
            "El mapa de calor integra en una sola visualización el promedio del trabajo escrito, la defensa, la nota final y la aprobación. Para permitir una lectura conjunta, la aprobación se transforma visualmente a una escala de 0 a 10, aunque la etiqueta conserva su porcentaje original.",
            heat,
            "El mapa permite localizar patrones relativos entre carreras sin depender de un solo indicador. Las áreas de menor desempeño deben verificarse con los tamaños de cohorte y con los criterios específicos de rúbrica antes de definir acciones.",
            height_cm=max(8.0, min(15.0, len(active_careers) * 0.55)),
        )

    if len(data["by_campus"]) > 1:
        campus_path = _chart_path(temp_paths)
        _bar_chart(
            [row["campus"] for row in data["by_campus"]],
            [float(row["records"]) for row in data["by_campus"]],
            "Registros PVC por sede", "Estudiantes", campus_path, horizontal=True
        )
        top_campus = max(data["by_campus"], key=lambda row: row["records"])
        _add_figure_block(
            body_story, context, styles, registry,
            "Distribución de registros PVC por sede",
            "Cuando la población conciliada corresponde a más de una sede, se presenta su distribución para contextualizar los resultados institucionales sin dividir el PVC en informes independientes.",
            campus_path,
            f"La sede con mayor número de registros conciliados es {top_campus['campus']} con {top_campus['records']} estudiantes. Las diferencias de volumen deben considerarse al comparar resultados entre sedes.",
        )

    # 6. Análisis estratégico
    pdf_progress_runtime._set_progress(70, "Construyendo análisis estratégico", "Generando Ishikawa, priorización y tendencias cuando los datos lo permiten.")
    _heading(body_story, context, styles, 1, "Análisis estratégico y mejora continua")
    ishikawa_path = _chart_path(temp_paths)
    _ishikawa(data, ishikawa_path)
    _add_figure_block(
        body_story, context, styles, registry,
        "Diagrama de Ishikawa de factores asociados y aspectos por verificar",
        "El Ishikawa organiza hallazgos observables del período en dimensiones de requisitos, trabajo escrito, defensa, calidad de datos, seguimiento y resultado. Las ramas representan factores asociados o aspectos por verificar; no constituyen una demostración causal.",
        ishikawa_path,
        f"Los datos registran {summary['failed']} reprobados entre evaluados, {summary['not_evaluated']} casos no evaluados, {summary['unmatched']} registros sin conciliación exacta y {summary['formula_warnings']} advertencias de fórmula. Estos hallazgos permiten orientar la revisión hacia evidencias concretas antes de formular decisiones.",
        note=SOURCE_LABEL + " El diagrama presenta asociaciones descriptivas, no relaciones causales demostradas.",
        height_cm=10.0,
    )

    priority_rows = [row for row in data["by_career"] if row["requirements"] and row["evaluated"]]
    if len(priority_rows) >= 2:
        priority = _chart_path(temp_paths)
        _priority_matrix(priority_rows, priority)
        _add_figure_block(
            body_story, context, styles, registry,
            "Matriz descriptiva de priorización para acciones de mejora",
            "La matriz combina dos indicadores observables por carrera: porcentaje de no habilitación en Requisitos y porcentaje de no aprobación entre estudiantes evaluados. Su uso es exclusivamente de priorización y no constituye un ranking de calidad académica.",
            priority,
            "Las carreras ubicadas más lejos del origen concentran simultáneamente mayores brechas de habilitación y/o aprobación. La priorización debe confirmarse con el número de estudiantes, los criterios de rúbrica y las causas documentadas de cada caso.",
        )

    history = _historical_pvc(report_id)
    if len(history) >= 2:
        history_path = _chart_path(temp_paths)
        _history_chart(history, history_path)
        current = next((row for row in history if row["current"]), None)
        _add_figure_block(
            body_story, context, styles, registry,
            "Evolución histórica de la aprobación PVC",
            "La comparación histórica se genera únicamente cuando Informtit dispone de al menos dos cohortes PVC con resultados evaluados. La línea muestra la tasa de aprobación calculada con el mismo denominador: estudiantes con evaluación final válida.",
            history_path,
            f"La cohorte actual registra una aprobación de {current['approval_pct']:.2f} %." if current else "La serie histórica permite observar la tendencia de aprobación entre cohortes disponibles.",
            note="Fuente: cohortes PVC almacenadas localmente en Informtit.",
        )

    # 7. Conclusiones
    _heading(body_story, context, styles, 1, "Conclusiones")
    _body(
        body_story, styles,
        f"La población maestra registra {summary['requirements_total']} estudiantes y {summary['eligible']} habilitados por requisitos. La Base PVC contiene {summary['pvc_total']} registros, de los cuales {summary['matched']} se conciliaron de manera exacta por cédula."
    )
    _body(
        body_story, styles,
        f"Entre {summary['evaluated']} estudiantes con evaluación final válida, {summary['approved']} aprobaron y {summary['failed']} reprobaron, con una tasa de aprobación de {summary['approval_pct']:.2f} %. Los {summary['not_evaluated']} casos sin evaluación se mantienen separados de los reprobados."
    )
    if practical_summary:
        weakest = min(practical_summary, key=lambda row: row["score"])
        _body(
            body_story, styles,
            f"En la evaluación práctica, el criterio con menor índice normalizado fue {weakest['criterion']} ({weakest['score']:.2f}/10), por lo que constituye un aspecto observable para el seguimiento metodológico."
        )
    if defense_summary:
        weakest = min(defense_summary, key=lambda row: row["score"])
        _body(
            body_story, styles,
            f"En la defensa, el criterio con menor índice normalizado fue {weakest['criterion']} ({weakest['score']:.2f}/10), dato que puede orientar actividades de preparación oral y argumentativa."
        )
    if summary["unmatched"] or summary["formula_warnings"]:
        _body(
            body_story, styles,
            f"Antes del cierre definitivo se recomienda resolver {summary['unmatched']} registros sin conciliación exacta y {summary['formula_warnings']} advertencias de fórmula, preservando la fuente oficial y la trazabilidad de cualquier corrección."
        )

    # 8. Plan de mejora
    _heading(body_story, context, styles, 1, "Plan de mejora")
    plan = []
    if summary["requirements_total"] - summary["eligible"] > 0:
        plan.append([
            "Requisitos habilitantes pendientes",
            f"{summary['requirements_total'] - summary['eligible']} estudiantes no constan habilitados",
            "Fortalecer el seguimiento anticipado de requisitos antes de la fase de evaluación.",
            "UTET y áreas responsables de requisitos",
            "Casos pendientes de requisitos",
            str(summary["requirements_total"] - summary["eligible"]),
            "Reducir respecto al período actual",
            "Próxima cohorte",
        ])
    if summary["unmatched"] > 0:
        plan.append([
            "Identidad sin conciliación",
            f"{summary['unmatched']} registros PVC",
            "Verificar cédula y corregir la fuente correspondiente antes del cierre.",
            "UTET",
            "Registros sin coincidencia",
            str(summary["unmatched"]),
            "0",
            "Antes del cierre",
        ])
    if summary["formula_warnings"] > 0:
        plan.append([
            "Diferencias de cálculo",
            f"{summary['formula_warnings']} advertencias",
            "Contrastar trabajo escrito, defensa y nota final con el acta o sistema fuente.",
            "UTET",
            "Advertencias de fórmula",
            str(summary["formula_warnings"]),
            "0",
            "Antes del cierre",
        ])
    if practical_summary:
        weakest = min(practical_summary, key=lambda row: row["score"])
        plan.append([
            f"Menor desempeño práctico: {weakest['criterion']}",
            f"Índice normalizado {weakest['score']:.2f}/10",
            "Reforzar el criterio en metodología, tutorías y retroalimentación del artículo.",
            "Docentes de metodología y tutores",
            f"Índice de {weakest['criterion']}",
            f"{weakest['score']:.2f}",
            "Mejorar respecto al período actual",
            "Próxima cohorte",
        ])
    if defense_summary:
        weakest = min(defense_summary, key=lambda row: row["score"])
        plan.append([
            f"Menor desempeño en defensa: {weakest['criterion']}",
            f"Índice normalizado {weakest['score']:.2f}/10",
            "Incorporar ejercicios de preparación oral y retroalimentación focalizada.",
            "Coordinaciones y docentes",
            f"Índice de {weakest['criterion']}",
            f"{weakest['score']:.2f}",
            "Mejorar respecto al período actual",
            "Próxima cohorte",
        ])
    if not plan:
        plan.append([
            "Seguimiento de mejora continua",
            "No se detectaron alertas críticas con los datos disponibles",
            "Mantener monitoreo de requisitos, evaluación y trazabilidad.",
            "UTET",
            "Indicadores PVC",
            "Período actual",
            "Mantener o mejorar",
            "Próxima cohorte",
        ])
    _add_table_block(
        body_story, context, styles, registry,
        "Matriz de acciones de mejora del proceso PVC",
        "Las acciones se derivan de hallazgos cuantificables del período. Cuando no existe una meta numérica institucional aprobada, la tabla evita inventarla y utiliza una meta comparativa o solicita reducir el indicador observado.",
        ["Hallazgo", "Evidencia", "Acción", "Responsable", "Indicador", "Actual", "Meta", "Plazo"],
        plan,
        [2.3 * cm, 2.3 * cm, 3.3 * cm, 2.5 * cm, 2.0 * cm, 1.2 * cm, 1.8 * cm, 1.5 * cm],
        "La matriz prioriza primero la integridad de datos y la habilitación, y luego los componentes académicos con menor desempeño relativo. Las metas no cuantificadas deben ser formalizadas por la instancia institucional competente para su seguimiento.",
        font_size=5.7,
    )

    # 9. Anexos nominales
    pdf_progress_runtime._set_progress(84, "Preparando anexos PVC", "Organizando calificaciones, tutorías y tribunal por estudiante.")
    _heading(body_story, context, styles, 1, "Anexos")
    individual_rows = [
        [
            row["career_name"] or "Sin conciliación",
            row["display_name"],
            row["identification"] or "—",
            _fmt(row["written"]),
            _fmt(row.get("defense_source")),
            _fmt(row["final_grade"]),
            row["final_status"],
        ]
        for row in data["records"]
    ]
    _add_table_block(
        body_story, context, styles, registry,
        "Detalle individual de resultados PVC",
        "El anexo nominal permite rastrear los indicadores agregados hasta el registro individual. La carrera y el nombre oficial proceden de Requisitos cuando existe conciliación exacta por cédula; los casos no conciliados se identifican expresamente.",
        ["Carrera", "Estudiante", "Cédula", "Escrito", "Defensa", "Final", "Estado"],
        individual_rows,
        [3.5 * cm, 4.6 * cm, 2.2 * cm, 1.5 * cm, 1.5 * cm, 1.4 * cm, 1.8 * cm],
        f"El anexo contiene {len(individual_rows)} registros y permite verificar las cifras consolidadas. Los registros sin coincidencia oficial no se asignan automáticamente a una carrera.",
        font_size=5.9,
    )
    roles_rows = [
        [
            row["display_name"],
            row.get("tutor_name") or "—",
            row.get("reader_name") or "—",
            " / ".join(filter(None, [row.get("vocal_1"), row.get("vocal_2"), row.get("vocal_3")])) or "—",
        ]
        for row in data["records"]
    ]
    _add_table_block(
        body_story, context, styles, registry,
        "Tutor, lector y tribunal registrados",
        "La siguiente tabla presenta los actores académicos vinculados a cada registro PVC, de acuerdo con la información consignada en la base de resultados.",
        ["Estudiante", "Tutor", "Lector", "Tribunal"],
        roles_rows,
        [4.1 * cm, 4.0 * cm, 4.0 * cm, 4.4 * cm],
        "La tabla constituye un control de trazabilidad de los actores registrados en la fuente PVC. Los campos vacíos deben verificarse en actas o registros institucionales antes del cierre.",
        font_size=5.6,
    )

    pdf_progress_runtime._set_progress(92, "Maquetando informe PVC", "Construyendo índices, encabezados y numeración institucional.")
    front: list[Any] = _cover(report, styles)
    front.append(Paragraph("ÍNDICE", styles["Title"]))
    toc = TableOfContents()
    toc.levelStyles = [
        ParagraphStyle("PVCTOC1", fontName="Helvetica-Bold", fontSize=10, leading=14, spaceBefore=4),
        ParagraphStyle("PVCTOC2", fontName="Helvetica", fontSize=9, leading=12, leftIndent=14),
        ParagraphStyle("PVCTOC3", fontName="Helvetica", fontSize=8, leading=11, leftIndent=28),
    ]
    front += [toc, PageBreak()]
    front.append(Paragraph("ÍNDICE DE TABLAS", styles["Title"]))
    for item in registry["tables"]:
        front.append(Paragraph(html.escape(item), styles["PvcSmall"]))
    front.append(PageBreak())
    front.append(Paragraph("ÍNDICE DE FIGURAS", styles["Title"]))
    for item in registry["figures"]:
        front.append(Paragraph(html.escape(item), styles["PvcSmall"]))
    front.append(PageBreak())

    story = front + body_story
    institutional.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    output = institutional.EXPORT_DIR / f"informtit_pvc_{report_id}.pdf"
    document = report_structure.TocDocTemplate(
        str(output),
        pagesize=A4,
        rightMargin=1.45 * cm,
        leftMargin=1.45 * cm,
        topMargin=3.4 * cm,
        bottomMargin=1.35 * cm,
        title=f"Informe PVC - {report.get('period', '')}",
    )
    try:
        document.multiBuild(
            story,
            canvasmaker=lambda *args, **kwargs: institutional.NumberedCanvas(
                *args, report=report, **kwargs
            ),
        )
    finally:
        for path in temp_paths:
            path.unlink(missing_ok=True)
    pdf_progress_runtime._set_progress(96, "Informe PVC maquetado", "Verificando el archivo final antes de la descarga.")
    return output


def _project_summary_pvc(project_id: int) -> dict[str, Any]:
    if _BASE_PROJECT_SUMMARY is None:
        return {}
    result = dict(_BASE_PROJECT_SUMMARY(project_id))
    if str(result.get("report_type") or "").lower() == "pvc":
        report_id = result.get("presencial_report_id") or result.get("id")
        result["pvc_records"] = _pvc_count(int(report_id) if report_id else None)
    return result


def install() -> None:
    global _INSTALLED, _BASE_BUILD_PDF, _BASE_PROJECT_SUMMARY
    if _INSTALLED:
        return
    ensure_schema()

    _BASE_BUILD_PDF = core.build_pdf
    _BASE_PROJECT_SUMMARY = period_unified_runtime._project_summary
    period_unified_runtime._project_summary = _project_summary_pvc

    def pdf_dispatch(report_id: int) -> Path:
        if _is_pvc(report_id):
            return build_pvc_pdf(report_id)
        if _BASE_BUILD_PDF is None:
            raise RuntimeError("Generador PDF base no disponible.")
        return Path(_BASE_BUILD_PDF(report_id))

    core.build_pdf = pdf_dispatch

    previous_get = core.InformtitHandler._handle_api_get
    previous_write = core.InformtitHandler._handle_api_write

    def pvc_get(self: Any, path: str, query: dict[str, list[str]]) -> None:
        audit_match = re.fullmatch(r"/api/reports/(\d+)/audit", path)
        if audit_match and _is_pvc(int(audit_match.group(1))):
            report_id = int(audit_match.group(1))
            audit = pvc_audit(report_id)
            token = pdf_progress_runtime.store_preflight(report_id, "pvc", audit)
            self._send_json({
                "ok": True,
                "audit": audit,
                "preflight_token": token,
            })
            return

        summary_match = re.fullmatch(r"/api/reports/(\d+)/pvc/summary", path)
        if summary_match:
            self._send_json(get_pvc_summary(int(summary_match.group(1)), include_records=True))
            return

        previous_get(self, path, query)

    def pvc_write(self: Any, method: str, path: str, payload: dict[str, Any]) -> None:
        import_match = re.fullmatch(r"/api/reports/(\d+)/pvc/import", path)
        if import_match and method == "POST":
            report_id = int(import_match.group(1))
            result = import_pvc_results(
                report_id,
                str(payload.get("data_url") or ""),
                str(payload.get("filename") or "base_pvc.xlsx"),
            )
            self._send_json(result, 201)
            return
        previous_write(self, method, path, payload)

    core.InformtitHandler._handle_api_get = pvc_get
    core.InformtitHandler._handle_api_write = pvc_write
    _INSTALLED = True
