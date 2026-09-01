from __future__ import annotations

import io
import unittest
from pathlib import Path

from openpyxl import Workbook

import pvc_report_runtime as pvc


class PvcReportRuntimeTests(unittest.TestCase):
    def workbook_bytes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "nombre_estudiante",
            "identificacion_estudiante",
            "periodo_academico",
            "trabajoTitulacion",
            "numeroActaGrado",
            "fechaActaGrado",
            "nombre_tutor",
            "nombre_lector",
            "nombre_vocal1",
            "nombre_vocal2",
            "nombre_vocal3",
            "evaluacionTutor",
            "evaluacionLector",
            "promedio_trabajo_escrito",
            "diseño_vocal1",
            "construccion_vocal1",
            "funcionamiento_vocal1",
            "aplicacion_vocal1",
            "sustentoMarcoTeorico_vocal1",
            "sustentoPropuesta_vocal1",
            "utilizacionRecursos_vocal1",
            "solvenciaPreguntas_vocal1",
            "diseño_vocal2",
            "construccion_vocal2",
            "funcionamiento_vocal2",
            "aplicacion_vocal2",
            "sustentoMarcoTeorico_vocal2",
            "sustentoPropuesta_vocal2",
            "utilizacionRecursos_vocal2",
            "solvenciaPreguntas_vocal2",
            "diseño_vocal3",
            "construccion_vocal3",
            "funcionamiento_vocal3",
            "aplicacion_vocal3",
            "sustentoMarcoTeorico_vocal3",
            "sustentoPropuesta_vocal3",
            "utilizacionRecursos_vocal3",
            "solvenciaPreguntas_vocal3",
            "promedio_trabajo_escrito",
            "notaPractico",
            "nota_defensa_oral",
            "notaTrabajoTitulacion",
            "notaPromedioAcumulado",
        ])
        sheet.append([
            "ESTUDIANTE DE PRUEBA", "1712345678", "Noviembre 2025 - Mayo 2026",
            "Artículo académico", "ACT-1", "2026-05-01", "Tutor", "Lector",
            "Vocal 1", "Vocal 2", "Vocal 3", 8.0, 8.0, 8.0,
            2.0, 2.0, 2.0, 2.0, 1.5, 1.5, 1.5, 3.0,
            2.0, 2.0, 2.0, 2.0, 1.5, 1.5, 1.5, 3.0,
            2.0, 2.0, 2.0, 2.0, 1.5, 1.5, 1.5, 3.0,
            8.0, "NULL", 7.5, 7.85, 8.2,
        ])
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def test_parser_accepts_real_pvc_shape_and_duplicate_written_header(self):
        rows = pvc.parse_pvc_workbook(self.workbook_bytes())
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["identification"], "1712345678")
        self.assertEqual(row["work_type"], "Artículo académico")
        self.assertEqual(row["written_calculated"], 8.0)
        self.assertEqual(row["final_calculated"], 7.85)
        self.assertEqual(row["final_source"], 7.85)
        self.assertEqual(row["formula_status"], "OK")
        self.assertEqual(len(row["practical"]), 4)
        self.assertEqual(len(row["defense"]), 4)

    def test_duplicate_identification_is_blocked_before_import(self):
        from openpyxl import load_workbook

        loaded = load_workbook(io.BytesIO(self.workbook_bytes()))
        sheet = loaded.active
        sheet.append([cell.value for cell in sheet[2]])
        stream = io.BytesIO()
        loaded.save(stream)

        with self.assertRaisesRegex(ValueError, "cédulas duplicadas"):
            pvc.parse_pvc_workbook(stream.getvalue())

    def test_pvc_formula_is_70_written_30_defense(self):
        self.assertEqual(pvc.WRITTEN_WEIGHT, 0.70)
        self.assertEqual(pvc.DEFENSE_WEIGHT, 0.30)
        calculated = round(8.0 * pvc.WRITTEN_WEIGHT + 7.5 * pvc.DEFENSE_WEIGHT, 2)
        self.assertEqual(calculated, 7.85)

    def test_missing_grade_is_not_automatically_failed(self):
        self.assertEqual(pvc._final_status(None, None, None), "NO EVALUADO")
        self.assertEqual(pvc._final_status(6.5, 7.0, 6.0), "REPROBADO")
        self.assertEqual(pvc._final_status(7.0, 7.0, 7.0), "APROBADO")

    def test_pdf_contract_enforces_context_before_and_analysis_after(self):
        source = Path("pvc_report_runtime.py").read_text(encoding="utf-8")
        table_start = source.index("def _add_table_block")
        figure_start = source.index("def _add_figure_block")
        table_chunk = source[table_start:figure_start]
        figure_chunk = source[figure_start:source.index("def _chart_path", figure_start)]
        self.assertLess(table_chunk.index("_body(story, styles, context_text)"), table_chunk.index("story.append(_table"))
        self.assertLess(table_chunk.index("story.append(_table"), table_chunk.index("_body(story, styles, analysis_text)"))
        self.assertLess(figure_chunk.index("_body(story, styles, context_text)"), figure_chunk.index("institutional.fit_image"))
        self.assertLess(figure_chunk.index("institutional.fit_image"), figure_chunk.index("_body(story, styles, analysis_text)"))

    def test_pvc_cover_uses_full_width_signature_grid(self):
        source = Path("pvc_report_runtime.py").read_text(encoding="utf-8")
        self.assertIn("colWidths=[6.0 * cm] * 3", source)
        self.assertIn("Spacer(1, 8.2 * cm)", source)

    def test_frontend_exposes_pvc_results_module(self):
        source = Path("static/pvc-report-ui.js").read_text(encoding="utf-8")
        generated = Path("static/generated-pdfs-ui.js").read_text(encoding="utf-8")
        self.assertIn("Resultados PVC", source)
        self.assertIn("PDF PVC", generated)
        self.assertIn("70 % trabajo escrito + 30 % defensa oral", source)
        self.assertIn("/pvc/import", source)
        self.assertIn("Regla del informe", source)
        runtime = Path("pvc_report_runtime.py").read_text(encoding="utf-8")
        self.assertIn("Modalidad Artículo Científico", runtime)
        self.assertNotIn("REQUISTOS", runtime)


if __name__ == "__main__":
    unittest.main()
