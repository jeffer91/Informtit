from __future__ import annotations

import csv
import io
import json
import re
import secrets
import unicodedata
import xml.etree.ElementTree as ET
from collections import Counter
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

import import_service


_INSTALLED = False
MAX_IMPORT_BYTES = 20 * 1024 * 1024
OLE_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_SIGNATURE = b"PK\x03\x04"


HEADER_ALIASES: dict[str, set[str]] = {
    "identification": {
        "numeroidentificacion", "numerodeidentificacion", "identificacion",
        "cedula", "cedulaidentidad", "ceduladeidentidad", "documento",
        "numerodocumento", "numerodedocumento", "ci",
    },
    "full_name": {
        "nombres", "nombre", "nombrecompleto", "nombresapellidos",
        "apellidosnombres", "estudiante", "nombreestudiante",
    },
    "career_code": {
        "codigocarrera", "codcarrera", "codigoprograma", "codprograma",
        "codigodelacarrera", "codigodecarrera",
    },
    "career_name": {
        "nombrecarrera", "carrera", "programa", "nombreprograma",
        "nombredecarrera", "programaacademico",
    },
    "schedule": {
        "horariocomplexivo", "horario", "jornada", "horarioexamen",
    },
    "academic_status": {"academico", "estadoacademico", "requisitoacademico"},
    "documentation_status": {
        "documentacion", "documentacionestado", "estadodocumentacion",
        "requisitodocumentacion",
    },
    "financial_status": {"financiero", "estadofinanciero", "requisitofinanciero"},
    "titulation_status": {"titulacion", "estadotitulacion", "requisitotitulacion"},
    "practices_linkage_status": {
        "practicasvinculacion", "practicasyvinculacion", "practicas",
        "practicaspreprofesionales",
    },
    "linkage_status": {"vinculacion", "estadovinculacion"},
    "graduate_followup_status": {
        "seguimientograduados", "seguimientoagraduados", "seguimiento",
    },
    "english_status": {"ingles", "estadoingles", "requisitoingles"},
    "data_update_status": {
        "actualizaciondatos", "actualizaciondedatos", "datosactualizados",
        "actualizardatos",
    },
    "personal_email": {
        "correopersonal", "emailpersonal", "mailpersonal", "correoprivado",
    },
    "email": {
        "correoinstitucional", "emailinstitucional", "mailinstitucional",
        "correoelectronico", "email", "correo",
    },
    "phone": {"celular", "telefono", "telefonocelular", "movil", "telefonoestudiante"},
    "campus": {"sede", "campus", "extension", "sedecampus"},
    "titulation_approval": {
        "aprobaciontitulacion", "aprobadoentitulacion", "resultadoaprobaciontitulacion",
    },
    "complexive_approval": {
        "aprobacioncomplexivoproyecto", "aprobacioncomplexivo", "aprobacionproyecto",
        "resultado complexivo", "resultadocomplexivo",
    },
}

REQUIRED_CANONICAL = {"identification", "full_name", "career_code", "career_name", "email"}


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", text.lower())


_ALIAS_TO_CANONICAL: dict[str, str] = {
    alias: canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _canonical_header(value: Any) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    exact = _ALIAS_TO_CANONICAL.get(normalized)
    if exact:
        return exact
    # Tolerancia para exportadores que agregan prefijos/sufijos descriptivos.
    candidates: list[tuple[int, str]] = []
    for alias, canonical in _ALIAS_TO_CANONICAL.items():
        if len(alias) < 7:
            continue
        if normalized.startswith(alias) or alias.startswith(normalized):
            candidates.append((len(alias), canonical))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return candidates[0][1]


class _HtmlRowsParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "tr":
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []
        elif tag == "br" and self._cell is not None:
            self._cell.append(" ")

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self._row is not None and self._cell is not None:
            self._row.append(import_service.clean_cell("".join(self._cell)))
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if any(import_service.clean_cell(value) for value in self._row):
                self.rows.append(self._row)
            self._row = None
            self._cell = None


def _decode_text(data: bytes) -> tuple[str, str]:
    attempts = (
        ("utf-8-sig", "UTF-8"),
        ("cp1252", "Windows-1252"),
        ("latin-1", "Latin-1"),
    )
    for encoding, label in attempts:
        try:
            return data.decode(encoding), label
        except UnicodeDecodeError:
            continue
    raise ValueError("No se pudo determinar la codificación del archivo de origen.")


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1].rsplit(":", 1)[-1]


def _rows_from_spreadsheetml(text: str) -> list[list[str]]:
    try:
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        raise ValueError("El XML de Excel no es válido.") from exc

    rows: list[list[str]] = []
    for row_node in root.iter():
        if _local_name(row_node.tag).lower() != "row":
            continue
        row: list[str] = []
        for cell in list(row_node):
            if _local_name(cell.tag).lower() != "cell":
                continue
            index_value = None
            for key, value in cell.attrib.items():
                if _local_name(key).lower() == "index":
                    try:
                        index_value = int(value)
                    except (TypeError, ValueError):
                        index_value = None
            if index_value and index_value > len(row) + 1:
                row.extend([""] * (index_value - len(row) - 1))
            data_node = next(
                (child for child in cell.iter() if _local_name(child.tag).lower() == "data"),
                None,
            )
            value = "" if data_node is None else "".join(data_node.itertext())
            row.append(import_service.clean_cell(value))
        if any(row):
            rows.append(row)
    if not rows:
        raise ValueError("El XML no contiene filas de hoja de cálculo.")
    return rows


def _rows_from_xlsx(data: bytes) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Falta el componente openpyxl para leer archivos .xlsx.") from exc

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("El archivo .xlsx no pudo abrirse correctamente.") from exc

    best_rows: list[list[Any]] = []
    best_score = -1
    for sheet in workbook.worksheets:
        rows = [["" if value is None else value for value in row] for row in sheet.iter_rows(values_only=True)]
        score = _header_score(rows)
        if score > best_score:
            best_score = score
            best_rows = rows
    workbook.close()
    return best_rows


def _rows_from_xls_binary(data: bytes) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Falta el componente xlrd para leer archivos .xls binarios antiguos.") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise ValueError("El archivo .xls binario no pudo abrirse correctamente.") from exc

    best_rows: list[list[Any]] = []
    best_score = -1
    try:
        for sheet in workbook.sheets():
            rows = [sheet.row_values(index) for index in range(sheet.nrows)]
            score = _header_score(rows)
            if score > best_score:
                best_score = score
                best_rows = rows
    finally:
        try:
            workbook.release_resources()
        except Exception:
            pass
    return best_rows


def _rows_from_delimited(text: str) -> tuple[list[list[str]], str]:
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    rows = [
        [import_service.clean_cell(value) for value in row]
        for row in csv.reader(io.StringIO(text), delimiter=delimiter)
        if any(import_service.clean_cell(value) for value in row)
    ]
    return rows, {",": "CSV", ";": "CSV (;)", "\t": "TSV", "|": "Texto delimitado (|)"}.get(delimiter, "CSV")


def _header_score(rows: Iterable[list[Any]]) -> int:
    best = 0
    for row in list(rows)[:40]:
        recognized = {_canonical_header(value) for value in row}
        recognized.discard(None)
        best = max(best, len(REQUIRED_CANONICAL.intersection(recognized)))
    return best


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[int, str], list[str]]:
    best_index = -1
    best_map: dict[int, str] = {}
    best_original: list[str] = []
    best_required = -1
    best_total = -1

    for index, row in enumerate(rows[:60]):
        mapping: dict[int, str] = {}
        seen: set[str] = set()
        for column, value in enumerate(row):
            canonical = _canonical_header(value)
            if canonical and canonical not in seen:
                mapping[column] = canonical
                seen.add(canonical)
        required = len(REQUIRED_CANONICAL.intersection(seen))
        total = len(seen)
        if (required, total) > (best_required, best_total):
            best_index = index
            best_map = mapping
            best_original = [import_service.clean_cell(value) for value in row]
            best_required = required
            best_total = total

    found = set(best_map.values())
    missing = sorted(REQUIRED_CANONICAL - found)
    if missing:
        friendly = {
            "identification": "identificación/cédula",
            "full_name": "nombres del estudiante",
            "career_code": "código de carrera",
            "career_name": "nombre de carrera",
            "email": "correo institucional",
        }
        labels = ", ".join(friendly[item] for item in missing)
        raise ValueError(
            "No se reconocieron todas las columnas mínimas del reporte. "
            f"Faltan: {labels}. Informtit reconoce nombres equivalentes y variaciones de mayúsculas, espacios y tildes."
        )
    return best_index, best_map, best_original


def _modality_details(career_name: str, career_code: str) -> tuple[str, str, str]:
    name = import_service.normalize_name(career_name)
    code = str(career_code or "").upper().strip()

    online_name = any(signal in name for signal in ("ONLINE", "EN LINEA", "VIRTUAL", "A DISTANCIA"))
    online_code = bool(re.search(r"(?:^|[-_/])L(?:[-_/]|$)", code)) or "-L-" in code
    presencial_name = "PRESENCIAL" in name
    presencial_code = bool(re.search(r"(?:^|[-_/])P(?:[-_/]|$)", code)) or "-P-" in code

    if online_name or online_code:
        reasons = []
        if online_name:
            reasons.append("nombre de carrera")
        if online_code:
            reasons.append("código -L-")
        return "en_linea", "alta", " + ".join(reasons)
    if presencial_name or presencial_code:
        reasons = []
        if presencial_name:
            reasons.append("nombre de carrera")
        if presencial_code:
            reasons.append("código -P-")
        return "presencial", "alta", " + ".join(reasons)

    inferred = import_service._modality(career_name, career_code)
    return inferred, "baja", "sin indicador explícito de modalidad"


def _records_from_rows(rows: list[list[Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    header_index, header_map, original_headers = _find_header(rows)
    records: list[dict[str, Any]] = []
    ambiguous: list[dict[str, str]] = []

    for source_row in rows[header_index + 1 :]:
        canonical: dict[str, str] = {}
        for column, key in header_map.items():
            value = source_row[column] if column < len(source_row) else ""
            canonical[key] = import_service.clean_cell(value)

        identification = canonical.get("identification", "")
        full_name = canonical.get("full_name", "")
        career_name = canonical.get("career_name", "")
        if not identification and not full_name:
            continue
        if not full_name or not career_name:
            continue

        career_code = canonical.get("career_code", "")
        modality, confidence, reason = _modality_details(career_name, career_code)
        if confidence == "baja":
            ambiguous.append(
                {
                    "identification": identification,
                    "student": full_name,
                    "career": career_name,
                    "career_code": career_code,
                    "inferred": modality,
                    "reason": reason,
                }
            )

        records.append(
            {
                "identification": identification,
                "full_name": full_name,
                "career_code": career_code,
                "career_name": career_name,
                "modality": modality,
                "modality_confidence": confidence,
                "modality_reason": reason,
                "schedule": canonical.get("schedule", ""),
                "academic_status": canonical.get("academic_status", ""),
                "documentation_status": canonical.get("documentation_status", ""),
                "financial_status": canonical.get("financial_status", ""),
                "titulation_status": canonical.get("titulation_status", ""),
                "practices_linkage_status": canonical.get("practices_linkage_status", ""),
                "linkage_status": canonical.get("linkage_status", ""),
                "graduate_followup_status": canonical.get("graduate_followup_status", ""),
                "english_status": canonical.get("english_status", ""),
                "data_update_status": canonical.get("data_update_status", ""),
                "personal_email": canonical.get("personal_email", ""),
                "email": canonical.get("email", "").lower(),
                "phone": canonical.get("phone", ""),
                "campus": canonical.get("campus", ""),
                "titulation_approval": canonical.get("titulation_approval", ""),
                "complexive_approval": canonical.get("complexive_approval", ""),
            }
        )

    if not records:
        raise ValueError("El archivo fue reconocido, pero no contiene estudiantes válidos después de normalizar las columnas.")

    meta = {
        "header_row": header_index + 1,
        "columns_recognized": len(set(header_map.values())),
        "columns_source": len(original_headers),
        "ambiguous_modality": len(ambiguous),
        "ambiguous_examples": ambiguous[:10],
    }
    return records, meta


def parse_roster_bytes(data: bytes, filename: str = "requisitos") -> dict[str, Any]:
    if not data:
        raise ValueError("El archivo está vacío.")
    if len(data) > MAX_IMPORT_BYTES:
        raise ValueError("El archivo supera el límite permitido de 20 MB.")

    stripped = data.lstrip()
    lower_probe = stripped[:1024 * 1024].lower()
    file_type = ""
    encoding = "Binario"
    rows: list[list[Any]]

    if data.startswith(OLE_SIGNATURE):
        file_type = "Excel 97-2003 binario (.xls)"
        rows = _rows_from_xls_binary(data)
    elif data.startswith(ZIP_SIGNATURE):
        file_type = "Excel moderno (.xlsx)"
        rows = _rows_from_xlsx(data)
    elif b"<table" in lower_probe:
        text, encoding = _decode_text(data)
        parser = _HtmlRowsParser()
        parser.feed(text)
        rows = parser.rows
        file_type = "HTML antiguo compatible con Excel"
    elif stripped.startswith(b"<"):
        text, encoding = _decode_text(data)
        rows = _rows_from_spreadsheetml(text)
        file_type = "Excel XML / SpreadsheetML"
    else:
        text, encoding = _decode_text(data)
        rows, file_type = _rows_from_delimited(text)

    if not rows:
        raise ValueError("No se encontraron filas tabulares en el archivo.")

    records, meta = _records_from_rows(rows)

    identification_counts = Counter(row["identification"] for row in records if row["identification"])
    email_counts = Counter(row["email"] for row in records if row["email"])
    career_counts: dict[str, Counter[str]] = {"presencial": Counter(), "en_linea": Counter()}
    for row in records:
        career_counts[row["modality"]][row["career_name"]] += 1

    presencial = sum(row["modality"] == "presencial" for row in records)
    online = sum(row["modality"] == "en_linea" for row in records)
    preview = {
        "filename": filename,
        "file_type": file_type,
        "encoding": encoding,
        "period": import_service._extract_period(filename),
        "total": len(records),
        "presencial": presencial,
        "en_linea": online,
        "careers_total": len({row["career_name"] for row in records}),
        "careers": {
            modality: [
                {"name": name, "students": count}
                for name, count in sorted(career_counts[modality].items())
            ]
            for modality in ("presencial", "en_linea")
        },
        "campuses": dict(sorted(Counter(row["campus"] for row in records if row["campus"]).items())),
        "schedules": dict(sorted(Counter(row["schedule"] for row in records if row["schedule"]).items())),
        "duplicate_identifications": [key for key, count in identification_counts.items() if count > 1],
        "duplicate_emails": [key for key, count in email_counts.items() if count > 1],
        "missing_institutional_email": sum(not row["email"] for row in records),
        **meta,
    }
    return {"records": records, "preview": preview}


def create_preview(data_url: str, filename: str) -> dict[str, Any]:
    data = import_service.decode_data_url(data_url)
    parsed = parse_roster_bytes(data, filename)
    token = secrets.token_urlsafe(18)
    base = import_service._import_dir() / token
    (base.with_suffix(".json")).write_text(json.dumps(parsed, ensure_ascii=False), encoding="utf-8")

    suffix = Path(filename).suffix.lower()
    if suffix not in {".xls", ".xlsx", ".csv", ".tsv", ".html", ".htm", ".xml"}:
        suffix = ".bin"
    (base.with_suffix(suffix)).write_bytes(data)
    return {"token": token, **parsed["preview"]}


def install() -> None:
    global _INSTALLED
    if _INSTALLED:
        return
    import_service.create_preview = create_preview
    import_service.parse_roster_bytes = parse_roster_bytes
    _INSTALLED = True
